import calendar
import math
import sqlite3
from collections import Counter
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import sys
from tkinter import BOTH, END, LEFT, RIGHT, VERTICAL, W, filedialog, messagebox, simpledialog
import tkinter as tk
from tkinter import ttk

VENDOR_DIR = Path(__file__).resolve().parent / 'vendor'
if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from rcs_module import RcsDashboardWindow


APP_TITLE = "Thuraya Prepay Airtime Sales"
PROJECT_DIR = Path(__file__).resolve().parent
DATA_DIR = PROJECT_DIR / "data"
ASSETS_DIR = PROJECT_DIR / "assets"
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "thuraya_airtime.sqlite3"
ALLOWED_VALUES = [20, 39, 50, 80, 160]
EXPECTED_HEADERS = [
    "serial number",
    "pin number",
    "value",
    "date of purchase",
    "expiry date",
    "supplier",
    "date of sale",
    "number of units sold in that sale",
    "client",
    "dealer",
    "invoice",
]
NAVY = "#0c2d48"
NAVY_LIGHT = "#123e64"
GREEN = "#1c8f78"
GREEN_LIGHT = "#2ab59a"
OFF_WHITE = "#f4f8fb"
CARD_BG = "#ffffff"
TEXT_DARK = "#183247"
MUTED = "#607587"
DOWNLOADS_DIR = Path(r"C:\Users\PASCA\OneDrive\Downloads")
PDF_HEADER_IMAGE = ASSETS_DIR / "RCSi header logo new.png"
PDF_FOOTER_IMAGE = ASSETS_DIR / "RCSi Footer.png"


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def normalize_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def normalize_date(value) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def digits_only(value) -> str:
    return "".join(ch for ch in str(value or "").strip() if ch.isdigit())


def safe_filename_part(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in (" ", "-", "_") else "_" for ch in value.strip())
    cleaned = " ".join(cleaned.split())
    return cleaned.replace(" ", "_") or "Client"


class CalendarPicker(tk.Toplevel):
    def __init__(self, parent, initial_date: str | None = None):
        super().__init__(parent)
        self.title("Choose Date")
        self.resizable(False, False)
        self.transient(parent)
        self.result: str | None = None
        parsed = normalize_date(initial_date) or date.today().isoformat()
        current = datetime.strptime(parsed, "%Y-%m-%d").date()
        self.displayed_year = current.year
        self.displayed_month = current.month
        self._build_ui()
        self._render_calendar()
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build_ui(self):
        header = ttk.Frame(self, padding=10)
        header.pack(fill="x")
        ttk.Button(header, text="<", width=3, command=self._prev_month).pack(side="left")
        self.month_label = ttk.Label(header, font=("Segoe UI Semibold", 11))
        self.month_label.pack(side="left", expand=True)
        ttk.Button(header, text=">", width=3, command=self._next_month).pack(side="right")

        days = ttk.Frame(self, padding=(10, 0, 10, 10))
        days.pack(fill="both", expand=True)
        for idx, name in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
            ttk.Label(days, text=name, anchor="center").grid(row=0, column=idx, padx=2, pady=(0, 6), sticky="nsew")
            days.columnconfigure(idx, weight=1)
        self.days_frame = days

    def _render_calendar(self):
        for child in list(self.days_frame.grid_slaves()):
            if int(child.grid_info().get("row", 0)) > 0:
                child.destroy()
        self.month_label.configure(text=f"{calendar.month_name[self.displayed_month]} {self.displayed_year}")
        month_rows = calendar.monthcalendar(self.displayed_year, self.displayed_month)
        for row_idx, week in enumerate(month_rows, start=1):
            for col_idx, day_num in enumerate(week):
                if day_num == 0:
                    ttk.Label(self.days_frame, text="").grid(row=row_idx, column=col_idx, padx=2, pady=2, sticky="nsew")
                    continue
                ttk.Button(
                    self.days_frame,
                    text=str(day_num),
                    width=4,
                    command=lambda d=day_num: self._select_day(d),
                ).grid(row=row_idx, column=col_idx, padx=2, pady=2, sticky="nsew")

    def _prev_month(self):
        if self.displayed_month == 1:
            self.displayed_year -= 1
            self.displayed_month = 12
        else:
            self.displayed_month -= 1
        self._render_calendar()

    def _next_month(self):
        if self.displayed_month == 12:
            self.displayed_year += 1
            self.displayed_month = 1
        else:
            self.displayed_month += 1
        self._render_calendar()

    def _select_day(self, day_num: int):
        self.result = date(self.displayed_year, self.displayed_month, day_num).isoformat()
        self.destroy()


def ask_date(parent, initial_date: str | None = None) -> str | None:
    picker = CalendarPicker(parent, initial_date=initial_date)
    parent.wait_window(picker)
    return picker.result


@dataclass
class SaleResult:
    sale_id: int
    sale_code: str
    excel_path: Path
    pdf_path: Path
    allocated_rows: list[sqlite3.Row]


class DatabaseManager:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self._ensure_database()

    @contextmanager
    def connect(self):
        connection = sqlite3.connect(self.db_path)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        try:
            yield connection
            connection.commit()
        finally:
            connection.close()

    def _ensure_database(self):
        with self.connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS sales (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sale_code TEXT NOT NULL UNIQUE,
                    client TEXT NOT NULL,
                    dealer TEXT NOT NULL,
                    invoice TEXT,
                    sale_date TEXT NOT NULL,
                    value INTEGER NOT NULL,
                    units_requested INTEGER NOT NULL,
                    units_allocated INTEGER NOT NULL,
                    output_xlsx TEXT NOT NULL,
                    output_pdf TEXT NOT NULL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS scratchcards (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    serial_number TEXT NOT NULL UNIQUE,
                    pin_number TEXT NOT NULL UNIQUE,
                    value INTEGER NOT NULL,
                    date_of_purchase TEXT,
                    expiry_date TEXT,
                    supplier TEXT NOT NULL DEFAULT 'Xtralink',
                    date_of_sale TEXT,
                    number_of_units_sold_in_sale INTEGER,
                    client TEXT,
                    dealer TEXT,
                    invoice TEXT,
                    sale_id INTEGER,
                    imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (sale_id) REFERENCES sales(id)
                );

                CREATE INDEX IF NOT EXISTS idx_scratchcards_value_unsold
                ON scratchcards(value, date_of_sale, expiry_date, date_of_purchase);

                CREATE INDEX IF NOT EXISTS idx_scratchcards_sale_date
                ON scratchcards(date_of_sale);
                """
            )

            sales_columns = {row["name"] for row in conn.execute("PRAGMA table_info(sales)").fetchall()}
            if "invoice" not in sales_columns:
                conn.execute("ALTER TABLE sales ADD COLUMN invoice TEXT")

            scratchcard_columns = {row["name"] for row in conn.execute("PRAGMA table_info(scratchcards)").fetchall()}
            if "invoice" not in scratchcard_columns:
                conn.execute("ALTER TABLE scratchcards ADD COLUMN invoice TEXT")

            conn.execute(
                """
                UPDATE scratchcards
                SET date_of_sale = COALESCE(date_of_sale, (SELECT sale_date FROM sales WHERE sales.id = scratchcards.sale_id)),
                    number_of_units_sold_in_sale = COALESCE(number_of_units_sold_in_sale, (SELECT units_allocated FROM sales WHERE sales.id = scratchcards.sale_id)),
                    client = COALESCE(client, (SELECT client FROM sales WHERE sales.id = scratchcards.sale_id)),
                    dealer = COALESCE(dealer, (SELECT dealer FROM sales WHERE sales.id = scratchcards.sale_id)),
                    invoice = COALESCE(invoice, (SELECT invoice FROM sales WHERE sales.id = scratchcards.sale_id))
                WHERE sale_id IS NOT NULL
                  AND (
                      date_of_sale IS NULL
                      OR number_of_units_sold_in_sale IS NULL
                      OR client IS NULL
                      OR dealer IS NULL
                      OR invoice IS NULL
                  )
                """
            )

    def import_scratchcards(self, file_path: Path) -> dict[str, object]:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(sheet[1], start=1):
            header_map[normalize_header(cell.value)] = idx

        missing = [header for header in EXPECTED_HEADERS if header not in header_map]
        if missing:
            raise ValueError(f"Missing columns in Excel file: {', '.join(missing)}")

        inserted = 0
        updated = 0
        inserted_by_value = Counter()
        updated_by_value = Counter()
        with self.connect() as conn:
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {
                    header: sheet.cell(row=row_idx, column=header_map[header]).value
                    for header in EXPECTED_HEADERS
                }
                if all(value in (None, "") for value in row_data.values()):
                    continue

                serial_number = digits_only(row_data["serial number"])
                pin_number = digits_only(row_data["pin number"])
                if len(serial_number) != 15:
                    raise ValueError(f"Row {row_idx}: serial number must be 15 digits.")
                if len(pin_number) > 14:
                    raise ValueError(f"Row {row_idx}: pin number must not exceed 14 digits.")
                pin_number = pin_number.zfill(14)

                try:
                    value = int(row_data["value"])
                except (TypeError, ValueError):
                    raise ValueError(f"Row {row_idx}: value must be one of {ALLOWED_VALUES}.")
                if value not in ALLOWED_VALUES:
                    raise ValueError(f"Row {row_idx}: value must be one of {ALLOWED_VALUES}.")

                supplier = str(row_data["supplier"]).strip() if row_data["supplier"] not in (None, "") else "Xtralink"
                date_of_purchase = normalize_date(row_data["date of purchase"])
                expiry_date = normalize_date(row_data["expiry date"])
                date_of_sale = normalize_date(row_data["date of sale"])
                units_sold = row_data["number of units sold in that sale"]
                units_sold = int(units_sold) if units_sold not in (None, "") else None
                client = str(row_data["client"]).strip() if row_data["client"] not in (None, "") else None
                dealer = str(row_data["dealer"]).strip() if row_data["dealer"] not in (None, "") else None
                invoice = str(row_data["invoice"]).strip() if row_data["invoice"] not in (None, "") else None

                existing = conn.execute(
                    "SELECT id FROM scratchcards WHERE serial_number = ?",
                    (serial_number,),
                ).fetchone()

                conn.execute(
                    """
                    INSERT INTO scratchcards (
                        serial_number, pin_number, value, date_of_purchase, expiry_date, supplier,
                        date_of_sale, number_of_units_sold_in_sale, client, dealer, invoice, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(serial_number) DO UPDATE SET
                        pin_number = excluded.pin_number,
                        value = excluded.value,
                        date_of_purchase = excluded.date_of_purchase,
                        expiry_date = excluded.expiry_date,
                        supplier = excluded.supplier,
                        date_of_sale = excluded.date_of_sale,
                        number_of_units_sold_in_sale = excluded.number_of_units_sold_in_sale,
                        client = excluded.client,
                        dealer = excluded.dealer,
                        invoice = excluded.invoice,
                        updated_at = excluded.updated_at
                    """,
                    (
                        serial_number,
                        pin_number,
                        value,
                        date_of_purchase,
                        expiry_date,
                        supplier,
                        date_of_sale,
                        units_sold,
                        client,
                        dealer,
                        invoice,
                        now_iso(),
                    ),
                )
                if existing:
                    updated += 1
                    updated_by_value[value] += 1
                else:
                    inserted += 1
                    inserted_by_value[value] += 1

        return {
            "inserted": inserted,
            "updated": updated,
            "inserted_by_value": dict(inserted_by_value),
            "updated_by_value": dict(updated_by_value),
        }

    def get_dashboard_metrics(self) -> dict[str, int]:
        with self.connect() as conn:
            totals = conn.execute(
                """
                SELECT
                    COUNT(*) AS total_cards,
                    SUM(CASE WHEN date_of_sale IS NULL THEN 1 ELSE 0 END) AS available_cards,
                    SUM(CASE WHEN date_of_sale IS NOT NULL THEN 1 ELSE 0 END) AS sold_cards,
                    COALESCE(SUM(CASE WHEN date_of_sale IS NULL THEN value ELSE 0 END), 0) AS available_value,
                    COALESCE(SUM(CASE WHEN date_of_sale IS NOT NULL THEN value ELSE 0 END), 0) AS sold_value
                FROM scratchcards
                """
            ).fetchone()
            sale_totals = conn.execute(
                """
                SELECT COUNT(*) AS sales_count,
                       COALESCE(SUM(units_allocated), 0) AS units_sold_total
                FROM sales
                """
            ).fetchone()

        return {
            "total_cards": totals["total_cards"] or 0,
            "available_cards": totals["available_cards"] or 0,
            "sold_cards": totals["sold_cards"] or 0,
            "available_value": totals["available_value"] or 0,
            "sold_value": totals["sold_value"] or 0,
            "sales_count": sale_totals["sales_count"] or 0,
            "units_sold_total": sale_totals["units_sold_total"] or 0,
        }

    def get_value_breakdown(self, sold: bool | None = None) -> list[sqlite3.Row]:
        query = """
            SELECT value, COUNT(*) AS qty, COALESCE(SUM(value), 0) AS total_value
            FROM scratchcards
        """
        params = []
        if sold is True:
            query += " WHERE date_of_sale IS NOT NULL"
        elif sold is False:
            query += " WHERE date_of_sale IS NULL"
        query += " GROUP BY value ORDER BY value"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def get_recent_sales_summary(self) -> list[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                """
                SELECT date_of_sale AS sale_date,
                       COUNT(DISTINCT COALESCE(NULLIF(TRIM(client), ''), '') || '|' || COALESCE(NULLIF(TRIM(dealer), ''), '') || '|' || COALESCE(NULLIF(TRIM(invoice), ''), '')) AS sale_count,
                       COUNT(*) AS units_sold,
                       COALESCE(SUM(value), 0) AS face_value_total
                FROM scratchcards
                WHERE date_of_sale IS NOT NULL
                GROUP BY date_of_sale
                ORDER BY date_of_sale DESC
                LIMIT 10
                """
            ).fetchall()

    def get_monthly_sales_summary(self) -> list[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                """
                WITH RECURSIVE month_span(month_start) AS (
                    SELECT date(MIN(date_of_sale), 'start of month')
                    FROM scratchcards
                    WHERE date_of_sale IS NOT NULL
                    UNION ALL
                    SELECT date(month_start, '+1 month')
                    FROM month_span
                    WHERE month_start < (
                        SELECT date(MAX(date_of_sale), 'start of month')
                        FROM scratchcards
                        WHERE date_of_sale IS NOT NULL
                    )
                ),
                monthly_sales AS (
                    SELECT
                        substr(date_of_sale, 1, 7) AS sale_month,
                        COUNT(*) AS units_sold
                    FROM scratchcards
                    WHERE date_of_sale IS NOT NULL
                    GROUP BY substr(date_of_sale, 1, 7)
                )
                SELECT
                    strftime('%Y-%m', month_span.month_start) AS sale_month,
                    COALESCE(monthly_sales.units_sold, 0) AS units_sold
                FROM month_span
                LEFT JOIN monthly_sales
                    ON monthly_sales.sale_month = strftime('%Y-%m', month_span.month_start)
                WHERE month_span.month_start IS NOT NULL
                ORDER BY month_span.month_start
                """
            ).fetchall()

    def list_inventory(self, value_filter: str = "All", status_filter: str = "All", search_text: str = "") -> list[sqlite3.Row]:
        conditions = []
        params = []
        if value_filter != "All":
            conditions.append("value = ?")
            params.append(int(value_filter))
        if status_filter == "Available":
            conditions.append("date_of_sale IS NULL")
        elif status_filter == "Sold":
            conditions.append("date_of_sale IS NOT NULL")
        if search_text:
            conditions.append(
                """
                (
                    serial_number LIKE ?
                    OR pin_number LIKE ?
                    OR IFNULL(client, '') LIKE ?
                    OR IFNULL(dealer, '') LIKE ?
                    OR IFNULL(invoice, '') LIKE ?
                    OR IFNULL(supplier, '') LIKE ?
                )
                """
            )
            wildcard = f"%{search_text.strip()}%"
            params.extend([wildcard, wildcard, wildcard, wildcard, wildcard, wildcard])

        query = "SELECT * FROM scratchcards"
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY value, date_of_sale IS NOT NULL, COALESCE(expiry_date, '9999-12-31'), serial_number"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def list_sales(self) -> list[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                """
                SELECT *
                FROM sales
                ORDER BY sale_date DESC, id DESC
                """
            ).fetchall()

    def list_history_groups(self, client_search: str = "", date_search: str = "", dealer_search: str = "", invoice_search: str = "") -> list[sqlite3.Row]:
        conditions = ["date_of_sale IS NOT NULL"]
        params: list[str] = []
        if client_search.strip():
            conditions.append("IFNULL(client, '') LIKE ?")
            params.append(f"%{client_search.strip()}%")
        if date_search.strip():
            conditions.append("IFNULL(date_of_sale, '') LIKE ?")
            params.append(f"%{date_search.strip()}%")
        if dealer_search.strip():
            conditions.append("IFNULL(dealer, '') LIKE ?")
            params.append(f"%{dealer_search.strip()}%")
        if invoice_search.strip():
            conditions.append("IFNULL(invoice, '') LIKE ?")
            params.append(f"%{invoice_search.strip()}%")

        query = f"""
            SELECT
                date_of_sale AS sale_date,
                IFNULL(client, '') AS client,
                IFNULL(dealer, '') AS dealer,
                IFNULL(invoice, '') AS invoice,
                COUNT(*) AS units_sold,
                COALESCE(SUM(value), 0) AS face_value_total,
                GROUP_CONCAT(DISTINCT CAST(value AS TEXT)) AS value_mix
            FROM scratchcards
            WHERE {' AND '.join(conditions)}
            GROUP BY date_of_sale, IFNULL(client, ''), IFNULL(dealer, ''), IFNULL(invoice, '')
            ORDER BY date_of_sale DESC, client, dealer, invoice
        """
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def get_history_group_rows(self, sale_date: str, client: str, dealer: str, invoice: str) -> list[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                """
                SELECT serial_number, pin_number, value, supplier, invoice, date_of_purchase, expiry_date
                FROM scratchcards
                WHERE IFNULL(date_of_sale, '') = ?
                  AND IFNULL(client, '') = ?
                  AND IFNULL(dealer, '') = ?
                  AND IFNULL(invoice, '') = ?
                ORDER BY value, serial_number
                """,
                (sale_date, client, dealer, invoice),
            ).fetchall()

    def available_count_for_value(self, value: int) -> int:
        with self.connect() as conn:
            row = conn.execute(
                "SELECT COUNT(*) AS qty FROM scratchcards WHERE value = ? AND date_of_sale IS NULL",
                (value,),
            ).fetchone()
        return row["qty"] or 0

    def stock_count_summary(self) -> list[sqlite3.Row]:
        with self.connect() as conn:
            return conn.execute(
                """
                SELECT value, COUNT(*) AS qty
                FROM scratchcards
                WHERE date_of_sale IS NULL
                  AND (client IS NULL OR TRIM(client) = '')
                  AND (dealer IS NULL OR TRIM(dealer) = '')
                GROUP BY value
                ORDER BY value
                """
            ).fetchall()

    def create_sale(self, client: str, dealer: str, invoice: str, sale_date: str, value: int, units: int) -> SaleResult:
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT *
                FROM scratchcards
                WHERE value = ? AND date_of_sale IS NULL
                ORDER BY COALESCE(expiry_date, '9999-12-31'), COALESCE(date_of_purchase, '9999-12-31'), serial_number
                LIMIT ?
                """,
                (value, units),
            ).fetchall()

            if len(rows) < units:
                raise ValueError(
                    f"Only {len(rows)} unsold scratchcards are available for value {value}. Requested {units}."
                )

            sale_date_iso = normalize_date(sale_date)
            if not sale_date_iso:
                raise ValueError("Sale date is required.")

            invoice_text = invoice.strip()
            filename_root = f"{sale_date_iso[2:4]}{sale_date_iso[5:7]}{sale_date_iso[8:10]}-{value}x{units}-{safe_filename_part(client)}"
            excel_path = DOWNLOADS_DIR / f"{filename_root}.xlsx"
            pdf_path = DOWNLOADS_DIR / f"{filename_root}.pdf"
            sale_code = filename_root

            cursor = conn.execute(
                """
                INSERT INTO sales (
                    sale_code, client, dealer, invoice, sale_date, value, units_requested,
                    units_allocated, output_xlsx, output_pdf
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sale_code,
                    client.strip(),
                    dealer.strip(),
                    invoice_text,
                    sale_date_iso,
                    value,
                    units,
                    units,
                    str(excel_path),
                    str(pdf_path),
                ),
            )
            sale_id = cursor.lastrowid

            row_ids = [row["id"] for row in rows]
            conn.executemany(
                """
                UPDATE scratchcards
                SET date_of_sale = ?, number_of_units_sold_in_sale = ?, client = ?, dealer = ?, invoice = ?, sale_id = ?, updated_at = ?
                WHERE id = ?
                """,
                [
                    (
                        sale_date_iso,
                        units,
                        client.strip(),
                        dealer.strip(),
                        invoice_text,
                        sale_id,
                        now_iso(),
                        row_id,
                    )
                    for row_id in row_ids
                ],
            )

        return SaleResult(sale_id=sale_id, sale_code=sale_code, excel_path=excel_path, pdf_path=pdf_path, allocated_rows=rows)

    def update_field(self, scratchcard_id: int, field_name: str, value: str | None):
        allowed_fields = {
            "date_of_purchase",
            "expiry_date",
            "supplier",
            "date_of_sale",
            "number_of_units_sold_in_sale",
            "client",
            "dealer",
            "invoice",
        }
        if field_name not in allowed_fields:
            raise ValueError("That field is not editable.")

        if field_name in {"date_of_purchase", "expiry_date", "date_of_sale"}:
            value = normalize_date(value)
        elif field_name == "number_of_units_sold_in_sale":
            value = int(value) if value not in (None, "") else None
        elif isinstance(value, str):
            value = value.strip() or None

        with self.connect() as conn:
            conn.execute(
                f"UPDATE scratchcards SET {field_name} = ?, updated_at = ? WHERE id = ?",
                (value, now_iso(), scratchcard_id),
            )


class ReportService:
    def _draw_pdf_branding(self, canvas, doc):
        page_width, page_height = A4
        usable_width = page_width - doc.leftMargin - doc.rightMargin
        if PDF_HEADER_IMAGE.exists():
            header = ImageReader(str(PDF_HEADER_IMAGE))
            header_width, header_height = header.getSize()
            target_width = usable_width * 0.33
            target_height = target_width * (header_height / header_width)
            canvas.drawImage(
                header,
                doc.leftMargin + usable_width - target_width,
                page_height - doc.topMargin + (2 * mm),
                width=target_width,
                height=target_height,
                preserveAspectRatio=True,
                mask="auto",
            )
        if PDF_FOOTER_IMAGE.exists():
            footer = ImageReader(str(PDF_FOOTER_IMAGE))
            footer_width, footer_height = footer.getSize()
            target_width = usable_width * 0.78
            target_height = target_width * (footer_height / footer_width)
            canvas.drawImage(
                footer,
                doc.leftMargin + ((usable_width - target_width) / 2),
                max(6 * mm, doc.bottomMargin - target_height - (2 * mm)),
                width=target_width,
                height=target_height,
                preserveAspectRatio=True,
                mask="auto",
            )

    def _build_export_paths(self, client: str, sale_date: str, rows: list[sqlite3.Row]) -> tuple[Path, Path]:
        unique_values = sorted({int(row["value"]) for row in rows})
        value_label = str(unique_values[0]) if len(unique_values) == 1 else "Mixed"
        filename_root = f"{sale_date[2:4]}{sale_date[5:7]}{sale_date[8:10]}-{value_label}x{len(rows)}-{safe_filename_part(client)}"
        return DOWNLOADS_DIR / f"{filename_root}.xlsx", DOWNLOADS_DIR / f"{filename_root}.pdf"

    def export_group_rows(self, rows: list[sqlite3.Row], client: str, dealer: str, sale_date: str) -> tuple[Path, Path]:
        if not rows:
            raise ValueError("No rows are available to export.")
        excel_path, pdf_path = self._build_export_paths(client, sale_date, rows)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sale"
        headers = ["serial number", "pin number", "client", "date", "value"]
        sheet.append(headers)
        for row in rows:
            sheet.append([
                f"'{row['serial_number']}",
                f"'{row['pin_number']}",
                client,
                sale_date,
                row["value"],
            ])
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
        for column, width in {"A": 20, "B": 20, "C": 24, "D": 14, "E": 10}.items():
            sheet.column_dimensions[column].width = width
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(excel_path)

        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=A4,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=32 * mm,
            bottomMargin=22 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("SaleTitle", parent=styles["Heading1"], textColor=NAVY, fontSize=16, spaceAfter=8)
        info_style = ParagraphStyle("InfoStyle", parent=styles["BodyText"], fontSize=10, textColor=TEXT_DARK, spaceAfter=3)
        table_data = [["serial number", "pin number", "client", "date", "value"]]
        for row in rows:
            table_data.append([row["serial_number"], row["pin_number"], client, sale_date, str(row["value"])])
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(NAVY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9cb7c8")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#edf6fb")]),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story = [
            Paragraph(APP_TITLE, title_style),
            Paragraph(f"Client: {client}", info_style),
            Paragraph(f"By: {dealer}", info_style),
            Paragraph(f"Sale date: {sale_date}", info_style),
            Paragraph(f"Units: {len(rows)}", info_style),
            Spacer(1, 6),
            table,
        ]
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        doc.build(story, onFirstPage=self._draw_pdf_branding, onLaterPages=self._draw_pdf_branding)
        return excel_path, pdf_path

    def create_excel_report(self, sale_result: SaleResult, client: str, sale_date: str, value: int):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sale"
        headers = ["serial number", "pin number", "client", "date", "value"]
        sheet.append(headers)
        for row in sale_result.allocated_rows:
            sheet.append(
                [
                    f"'{row['serial_number']}",
                    f"'{row['pin_number']}",
                    client,
                    sale_date,
                    value,
                ]
            )

        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)

        widths = {"A": 20, "B": 20, "C": 24, "D": 14, "E": 10}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        sale_result.excel_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(sale_result.excel_path)

    def create_pdf_report(self, sale_result: SaleResult, client: str, dealer: str, sale_date: str, value: int):
        sale_result.pdf_path.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(
            str(sale_result.pdf_path),
            pagesize=A4,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=32 * mm,
            bottomMargin=22 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "SaleTitle",
            parent=styles["Heading1"],
            textColor=NAVY,
            fontSize=16,
            spaceAfter=8,
        )
        info_style = ParagraphStyle(
            "InfoStyle",
            parent=styles["BodyText"],
            fontSize=10,
            textColor=TEXT_DARK,
            spaceAfter=3,
        )

        table_data = [["serial number", "pin number", "client", "date", "value"]]
        for row in sale_result.allocated_rows:
            table_data.append(
                [
                    row["serial_number"],
                    row["pin_number"],
                    client,
                    sale_date,
                    str(value),
                ]
            )

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(NAVY)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9cb7c8")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#edf6fb")]),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story = [
            Paragraph(APP_TITLE, title_style),
            Paragraph(f"Client: {client}", info_style),
            Paragraph(f"By: {dealer}", info_style),
            Paragraph(f"Sale date: {sale_date}", info_style),
            Paragraph(f"Card value: {value}", info_style),
            Paragraph(f"Units: {len(sale_result.allocated_rows)}", info_style),
            Spacer(1, 6),
            table,
        ]
        doc.build(story, onFirstPage=self._draw_pdf_branding, onLaterPages=self._draw_pdf_branding)


class KPIStat(ttk.Frame):
    def __init__(self, parent, title: str, value: str, accent: str):
        super().__init__(parent, style="Card.TFrame", padding=14)
        self.columnconfigure(0, weight=1)
        ttk.Label(self, text=title, style="CardTitle.TLabel").grid(row=0, column=0, sticky="w")
        self.value_label = ttk.Label(self, text=value, style="CardValue.TLabel", foreground=accent)
        self.value_label.grid(row=1, column=0, sticky="w", pady=(6, 0))

    def set_value(self, value: str):
        self.value_label.configure(text=value)


class ThurayaApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1450x920")
        self.minsize(1280, 820)
        self.configure(bg=OFF_WHITE)

        self.db = DatabaseManager(DB_PATH)
        self.reports = ReportService()
        self.selected_inventory_row_id: int | None = None
        self.inventory_sort_column = "date_of_sale"
        self.inventory_sort_reverse = True
        self.selected_history_group: tuple[str, str, str, str] | None = None
        self.inventory_results_var = tk.StringVar(value="Result of search: 0")
        self.history_client_search_var = tk.StringVar()
        self.history_date_search_var = tk.StringVar()
        self.history_dealer_search_var = tk.StringVar()
        self.history_invoice_search_var = tk.StringVar()
        self.history_detail_var = tk.StringVar(value="Click a history row to see the sale details.")
        self.rcs_dashboard_window = None

        self._configure_style()
        self._build_header()
        self._build_notebook()
        self.refresh_all()

    def _configure_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(".", font=("Segoe UI", 10))
        style.configure("App.TFrame", background=OFF_WHITE)
        style.configure("Card.TFrame", background=CARD_BG, relief="flat")
        style.configure("Header.TFrame", background=NAVY)
        style.configure("HeroTitle.TLabel", background=NAVY, foreground="white", font=("Segoe UI Semibold", 20))
        style.configure("HeroSub.TLabel", background=NAVY, foreground="#d8e7f0", font=("Segoe UI", 10))
        style.configure("SectionTitle.TLabel", background=OFF_WHITE, foreground=TEXT_DARK, font=("Segoe UI Semibold", 14))
        style.configure("CardTitle.TLabel", background=CARD_BG, foreground=MUTED, font=("Segoe UI", 10))
        style.configure("CardValue.TLabel", background=CARD_BG, foreground=GREEN, font=("Segoe UI Semibold", 20))
        style.configure("TLabel", background=OFF_WHITE, foreground=TEXT_DARK)
        style.configure("TNotebook", background=OFF_WHITE, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(16, 10), background="#d9e7ef", foreground=TEXT_DARK)
        style.map("TNotebook.Tab", background=[("selected", CARD_BG)], foreground=[("selected", NAVY)])
        style.configure("Primary.TButton", background=GREEN, foreground="white", padding=(12, 8), borderwidth=0)
        style.map(
            "Primary.TButton",
            background=[("active", GREEN_LIGHT), ("pressed", GREEN_LIGHT)],
            foreground=[("disabled", "#dbe5eb"), ("!disabled", "white")],
        )
        style.configure("Secondary.TButton", background=NAVY_LIGHT, foreground="white", padding=(12, 8), borderwidth=0)
        style.map("Secondary.TButton", background=[("active", NAVY), ("pressed", NAVY)])
        style.configure("Alert.TButton", background="#c0392b", foreground="white", padding=(12, 8), borderwidth=0)
        style.map("Alert.TButton", background=[("active", "#a93226"), ("pressed", "#a93226")])
        style.configure("Treeview", rowheight=28, background="white", fieldbackground="white", foreground=TEXT_DARK)
        style.configure("Treeview.Heading", background=NAVY, foreground="white", font=("Segoe UI Semibold", 10))
        style.map("Treeview", background=[("selected", "#d7efe8")], foreground=[("selected", TEXT_DARK)])
        style.configure("TLabelframe", background=OFF_WHITE, foreground=TEXT_DARK)
        style.configure("TLabelframe.Label", background=OFF_WHITE, foreground=TEXT_DARK, font=("Segoe UI Semibold", 11))
        style.configure("Status.TLabel", background=OFF_WHITE, foreground=MUTED)

    def _build_header(self):
        header = ttk.Frame(self, style="Header.TFrame", padding=(24, 18))
        header.pack(fill="x")
        ttk.Label(header, text=APP_TITLE, style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(
            header,
            text="Import scratchcards, allocate airtime sales, generate delivery files, and track performance in one Windows-friendly desktop app.",
            style="HeroSub.TLabel",
            wraplength=1050,
        ).pack(anchor="w", pady=(4, 0))

    def _build_notebook(self):
        body = ttk.Frame(self, style="App.TFrame", padding=18)
        body.pack(fill=BOTH, expand=True)

        self.launcher_frame = ttk.Frame(body, style="App.TFrame")
        self.launcher_frame.pack(fill=BOTH, expand=True)

        launcher_card = ttk.Frame(self.launcher_frame, style="Card.TFrame", padding=28)
        launcher_card.pack(expand=True)
        ttk.Label(launcher_card, text="Choose Workspace", style="SectionTitle.TLabel").pack(anchor="center", pady=(0, 10))
        ttk.Label(
            launcher_card,
            text="Open the Thuraya airtime workflow or the new RCS finance tools from one starting screen.",
            style="Status.TLabel",
            wraplength=520,
            justify="center",
        ).pack(anchor="center", pady=(0, 20))
        ttk.Button(launcher_card, text="RCS", style="Secondary.TButton", command=self.open_rcs_dashboard).pack(fill="x", pady=(0, 12))
        ttk.Button(launcher_card, text="Thuraya Airtime Sales", style="Primary.TButton", command=self.show_thuraya_screen).pack(fill="x")

        self.thuraya_shell = ttk.Frame(body, style="App.TFrame")
        top_bar = ttk.Frame(self.thuraya_shell, style="App.TFrame")
        top_bar.pack(fill="x", pady=(0, 10))
        ttk.Button(top_bar, text="Back to Main Menu", style="Secondary.TButton", command=self.show_home_screen).pack(anchor="w")

        self.notebook = ttk.Notebook(self.thuraya_shell)
        self.notebook.pack(fill=BOTH, expand=True)

        self.dashboard_tab = ttk.Frame(self.notebook, style="App.TFrame", padding=14)
        self.import_tab = ttk.Frame(self.notebook, style="App.TFrame", padding=14)
        self.sales_tab = ttk.Frame(self.notebook, style="App.TFrame", padding=14)
        self.inventory_tab = ttk.Frame(self.notebook, style="App.TFrame", padding=14)
        self.history_tab = ttk.Frame(self.notebook, style="App.TFrame", padding=14)
        self.analytics_tab = ttk.Frame(self.notebook, style="App.TFrame", padding=14)

        self.notebook.add(self.dashboard_tab, text="Dashboard")
        self.notebook.add(self.import_tab, text="Import")
        self.notebook.add(self.sales_tab, text="Sales")
        self.notebook.add(self.inventory_tab, text="Inventory")
        self.notebook.add(self.history_tab, text="History")
        self.notebook.add(self.analytics_tab, text="Analytics")

        self._build_dashboard_tab()
        self._build_import_tab()
        self._build_sales_tab()
        self._build_inventory_tab()
        self._build_history_tab()
        self._build_analytics_tab()

        self.rcs_shell = ttk.Frame(body, style="App.TFrame")

    def show_home_screen(self):
        self.thuraya_shell.pack_forget()
        self.rcs_shell.pack_forget()
        self.launcher_frame.pack(fill=BOTH, expand=True)

    def show_thuraya_screen(self):
        self.launcher_frame.pack_forget()
        self.rcs_shell.pack_forget()
        self.thuraya_shell.pack(fill=BOTH, expand=True)
        self.notebook.select(self.dashboard_tab)

    def open_rcs_dashboard(self):
        self.launcher_frame.pack_forget()
        self.thuraya_shell.pack_forget()
        if self.rcs_dashboard_window is None or not self.rcs_dashboard_window.winfo_exists():
            self.rcs_dashboard_window = RcsDashboardWindow(self.rcs_shell, DB_PATH, on_back=self.show_home_screen)
        self.rcs_dashboard_window.pack(fill=BOTH, expand=True)
        self.rcs_dashboard_window.show_home()
        self.rcs_shell.pack(fill=BOTH, expand=True)

    def _build_dashboard_tab(self):
        ttk.Label(self.dashboard_tab, text="Business Overview", style="SectionTitle.TLabel").pack(anchor="w", pady=(0, 12))

        stats_frame = ttk.Frame(self.dashboard_tab, style="App.TFrame")
        stats_frame.pack(fill="x")
        stats_frame.columnconfigure((0, 1, 2, 3), weight=1)

        self.stat_total_cards = KPIStat(stats_frame, "Total Cards", "0", NAVY)
        self.stat_available = KPIStat(stats_frame, "Available Cards", "0", GREEN)
        self.stat_sold = KPIStat(stats_frame, "Sold Cards", "0", NAVY_LIGHT)
        self.stat_sales_count = KPIStat(stats_frame, "Sales Count", "0", GREEN_LIGHT)

        for idx, widget in enumerate((self.stat_total_cards, self.stat_available, self.stat_sold, self.stat_sales_count)):
            widget.grid(row=0, column=idx, sticky="nsew", padx=(0 if idx == 0 else 10, 0))

        lower = ttk.Frame(self.dashboard_tab, style="App.TFrame")
        lower.pack(fill=BOTH, expand=True, pady=(14, 0))
        lower.columnconfigure(0, weight=3)
        lower.columnconfigure(1, weight=2)
        lower.rowconfigure(0, weight=1)

        left_card = ttk.Frame(lower, style="Card.TFrame", padding=14)
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        ttk.Label(left_card, text="Recent Sales Snapshot", style="SectionTitle.TLabel").pack(anchor="w")
        self.dashboard_sales_tree = self._create_tree(
            left_card,
            columns=("sale_date", "sale_count", "units_sold", "face_value_total"),
            headings=("Sale Date", "Sales", "Units", "Face Value"),
            height=14,
        )
        self.dashboard_sales_tree.pack(fill=BOTH, expand=True, pady=(10, 0))

        right_card = ttk.Frame(lower, style="Card.TFrame", padding=14)
        right_card.grid(row=0, column=1, sticky="nsew")
        ttk.Label(right_card, text="Quick Status", style="SectionTitle.TLabel").pack(anchor="w")
        self.stock_count_button = ttk.Button(right_card, text="Stock Count", style="Secondary.TButton", command=self.show_stock_count)
        self.stock_count_button.pack(anchor="w", pady=(10, 0))
        self.quick_status_text = tk.Text(
            right_card,
            height=14,
            relief="flat",
            bg="white",
            fg=TEXT_DARK,
            font=("Segoe UI", 10),
            wrap="word",
        )
        self.quick_status_text.pack(fill=BOTH, expand=True, pady=(10, 0))
        self.quick_status_text.configure(state="disabled")

    def _build_import_tab(self):
        top = ttk.Frame(self.import_tab, style="App.TFrame")
        top.pack(fill="x")
        top.columnconfigure(0, weight=1)

        left = ttk.Frame(top, style="Card.TFrame", padding=16)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        ttk.Label(left, text="Import Scratchcards from Excel", style="SectionTitle.TLabel").grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(
            left,
            text="Choose an .xlsx file that uses the Table 1 field names as column headers. Existing serial numbers are updated, new ones are inserted.",
            style="Status.TLabel",
            wraplength=620,
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 16))
        ttk.Label(left, text="Excel file").grid(row=2, column=0, sticky="w", pady=4)
        self.import_path_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.import_path_var, width=70).grid(row=3, column=0, sticky="ew", pady=(0, 12))
        ttk.Button(left, text="Browse", style="Secondary.TButton", command=self.browse_import_file).grid(row=3, column=1, padx=10, sticky="ew")
        ttk.Button(left, text="Import Now", style="Primary.TButton", command=self.import_excel).grid(row=3, column=2, sticky="ew")
        left.columnconfigure(0, weight=1)

        self.import_status_var = tk.StringVar(value="No import has been run yet.")
        ttk.Label(left, textvariable=self.import_status_var, style="Status.TLabel", wraplength=620).grid(row=4, column=0, columnspan=3, sticky="w", pady=(14, 0))

        right = ttk.Frame(top, style="Card.TFrame", padding=16)
        right.grid(row=0, column=1, sticky="nsew")
        ttk.Label(right, text="Expected Columns", style="SectionTitle.TLabel").pack(anchor="w")
        columns_box = tk.Text(right, height=14, relief="flat", bg="white", fg=TEXT_DARK, font=("Consolas", 10))
        columns_box.pack(fill=BOTH, expand=True, pady=(10, 0))
        columns_box.insert("1.0", "\n".join(EXPECTED_HEADERS))
        columns_box.configure(state="disabled")

    def _build_sales_tab(self):
        wrapper = ttk.Frame(self.sales_tab, style="App.TFrame")
        wrapper.pack(fill=BOTH, expand=True)
        wrapper.columnconfigure(0, weight=2)
        wrapper.columnconfigure(1, weight=3)
        wrapper.rowconfigure(0, weight=1)

        form_card = ttk.Frame(wrapper, style="Card.TFrame", padding=16)
        form_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        ttk.Label(form_card, text="Create Sale", style="SectionTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w")

        self.sale_client_var = tk.StringVar()
        self.sale_dealer_var = tk.StringVar()
        self.sale_invoice_var = tk.StringVar()
        self.sale_date_var = tk.StringVar(value=date.today().isoformat())
        self.sale_value_var = tk.StringVar(value=str(ALLOWED_VALUES[0]))
        self.sale_units_var = tk.StringVar()
        self.sale_available_var = tk.StringVar(value="Available cards for selected value: 0")

        fields = [
            ("Client name", self.sale_client_var),
            ("Dealer", self.sale_dealer_var),
            ("Invoice", self.sale_invoice_var),
            ("Selling date", self.sale_date_var),
            ("Number of units needed", self.sale_units_var),
        ]
        row = 1
        for label, var in fields:
            ttk.Label(form_card, text=label).grid(row=row, column=0, sticky="w", pady=(12, 4))
            ttk.Entry(form_card, textvariable=var).grid(row=row + 1, column=0, columnspan=2, sticky="ew")
            row += 2

        ttk.Label(form_card, text="Value needed").grid(row=row, column=0, sticky="w", pady=(12, 4))
        value_combo = ttk.Combobox(form_card, values=[str(v) for v in ALLOWED_VALUES], textvariable=self.sale_value_var, state="readonly")
        value_combo.grid(row=row + 1, column=0, columnspan=2, sticky="ew")
        value_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_sale_available())
        row += 2

        ttk.Label(form_card, textvariable=self.sale_available_var, style="Status.TLabel").grid(row=row, column=0, columnspan=2, sticky="w", pady=(12, 12))
        row += 1
        ttk.Button(form_card, text="Create Sale and Generate Files", style="Primary.TButton", command=self.run_sale).grid(row=row, column=0, columnspan=2, sticky="ew")
        row += 1
        self.sale_status_var = tk.StringVar(value="No sale created yet.")
        ttk.Label(form_card, textvariable=self.sale_status_var, style="Status.TLabel", wraplength=420).grid(row=row, column=0, columnspan=2, sticky="w", pady=(14, 0))
        form_card.columnconfigure((0, 1), weight=1)

        right_card = ttk.Frame(wrapper, style="Card.TFrame", padding=16)
        right_card.grid(row=0, column=1, sticky="nsew")
        ttk.Label(right_card, text="Latest Sales", style="SectionTitle.TLabel").pack(anchor="w")
        self.sales_tree = self._create_tree(
            right_card,
            columns=("sale_code", "sale_date", "client", "dealer", "invoice", "value", "units_allocated"),
            headings=("Sale Code", "Date", "Client", "Dealer", "Invoice", "Value", "Units"),
            height=18,
        )
        self.sales_tree.pack(fill=BOTH, expand=True, pady=(10, 0))

    def _build_inventory_tab(self):
        controls = ttk.Frame(self.inventory_tab, style="App.TFrame")
        controls.pack(fill="x")
        controls.columnconfigure(5, weight=1)

        self.inventory_value_var = tk.StringVar(value="All")
        self.inventory_status_var = tk.StringVar(value="All")
        self.inventory_search_var = tk.StringVar()

        ttk.Label(controls, text="Value").grid(row=0, column=0, sticky="w")
        value_combo = ttk.Combobox(controls, values=["All"] + [str(v) for v in ALLOWED_VALUES], textvariable=self.inventory_value_var, width=10, state="readonly")
        value_combo.grid(row=1, column=0, sticky="w", padx=(0, 10))
        value_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_inventory())

        ttk.Label(controls, text="Status").grid(row=0, column=1, sticky="w")
        status_combo = ttk.Combobox(controls, values=["All", "Available", "Sold"], textvariable=self.inventory_status_var, width=12, state="readonly")
        status_combo.grid(row=1, column=1, sticky="w", padx=(0, 10))
        status_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_inventory())

        ttk.Label(controls, text="Search").grid(row=0, column=2, sticky="w")
        search_entry = ttk.Entry(controls, textvariable=self.inventory_search_var, width=30)
        search_entry.grid(row=1, column=2, sticky="ew", padx=(0, 10))
        search_entry.bind("<Return>", lambda _event: self.refresh_inventory())

        ttk.Button(controls, text="Apply", style="Secondary.TButton", command=self.refresh_inventory).grid(row=1, column=3, padx=(0, 10))
        ttk.Button(controls, text="Edit Selected", style="Primary.TButton", command=self.edit_selected_inventory).grid(row=1, column=4)
        ttk.Label(controls, textvariable=self.inventory_results_var, style="Status.TLabel").grid(row=1, column=5, sticky="w", padx=(12, 0))

        card = ttk.Frame(self.inventory_tab, style="Card.TFrame", padding=12)
        card.pack(fill=BOTH, expand=True, pady=(12, 0))
        self.inventory_tree = self._create_tree(
            card,
            columns=(
                "id",
                "serial_number",
                "pin_number",
                "value",
                "date_of_purchase",
                "expiry_date",
                "supplier",
                "date_of_sale",
                "number_of_units_sold_in_sale",
                "client",
                "dealer",
                "invoice",
            ),
            headings=("ID", "Serial Number", "Pin Number", "Value", "Purchase Date", "Expiry Date", "Supplier", "Sale Date", "Units Sold", "Client", "Dealer", "Invoice"),
            height=20,
        )
        self.inventory_tree.pack(fill=BOTH, expand=True)
        inventory_widget = self.inventory_tree.children["!treeview"]
        inventory_widget.bind("<<TreeviewSelect>>", self.on_inventory_select)
        self._configure_inventory_sorting(inventory_widget)

    def _build_history_tab(self):
        controls = ttk.Frame(self.history_tab, style="App.TFrame")
        controls.pack(fill="x")
        ttk.Label(controls, text="Client").grid(row=0, column=0, sticky="w")
        ttk.Entry(controls, textvariable=self.history_client_search_var, width=24).grid(row=1, column=0, sticky="ew", padx=(0, 10))
        ttk.Label(controls, text="Date").grid(row=0, column=1, sticky="w")
        ttk.Entry(controls, textvariable=self.history_date_search_var, width=16).grid(row=1, column=1, sticky="ew", padx=(0, 6))
        ttk.Button(controls, text="Calendar", style="Secondary.TButton", command=self.pick_history_date).grid(row=1, column=2, padx=(0, 10))
        ttk.Label(controls, text="Dealer").grid(row=0, column=3, sticky="w")
        ttk.Entry(controls, textvariable=self.history_dealer_search_var, width=20).grid(row=1, column=3, sticky="ew", padx=(0, 10))
        ttk.Label(controls, text="Invoice").grid(row=0, column=4, sticky="w")
        ttk.Entry(controls, textvariable=self.history_invoice_search_var, width=20).grid(row=1, column=4, sticky="ew", padx=(0, 10))
        ttk.Button(controls, text="Search", style="Secondary.TButton", command=self.refresh_history).grid(row=1, column=5)

        upper_card = ttk.Frame(self.history_tab, style="Card.TFrame", padding=12)
        upper_card.pack(fill=BOTH, expand=True, pady=(12, 8))
        ttk.Label(upper_card, text="Sales History", style="SectionTitle.TLabel").pack(anchor="w", pady=(0, 10))
        self.history_tree = self._create_tree(
            upper_card,
            columns=("sale_date", "client", "dealer", "invoice", "units_sold", "face_value_total", "value_mix"),
            headings=("Date", "Client", "Dealer", "Invoice", "Units Sold", "Face Value", "Values Sold"),
            height=10,
        )
        self.history_tree.pack(fill=BOTH, expand=True)
        self.history_tree.children["!treeview"].bind("<<TreeviewSelect>>", self.on_history_select)

        detail_card = ttk.Frame(self.history_tab, style="Card.TFrame", padding=12)
        detail_card.pack(fill=BOTH, expand=True)
        detail_top = ttk.Frame(detail_card, style="Card.TFrame")
        detail_top.pack(fill="x")
        ttk.Label(detail_top, text="Sale Details", style="SectionTitle.TLabel").pack(anchor="w", side="left")
        ttk.Button(detail_top, text="Save Detail XLSX + PDF", style="Primary.TButton", command=self.export_selected_history_sale).pack(side="right")
        ttk.Label(detail_card, textvariable=self.history_detail_var, style="Status.TLabel").pack(anchor="w", pady=(6, 10))
        self.history_detail_tree = self._create_tree(
            detail_card,
            columns=("serial_number", "pin_number", "value", "invoice", "supplier", "date_of_purchase", "expiry_date"),
            headings=("Serial Number", "Pin Number", "Value", "Invoice", "Supplier", "Purchase Date", "Expiry Date"),
            height=10,
        )
        self.history_detail_tree.pack(fill=BOTH, expand=True)

    def _build_analytics_tab(self):
        outer = ttk.Frame(self.analytics_tab, style="App.TFrame")
        outer.pack(fill=BOTH, expand=True)
        outer.columnconfigure((0, 1), weight=1)
        outer.rowconfigure(0, weight=2)
        outer.rowconfigure(1, weight=1)

        self.sales_chart_card = ttk.Frame(outer, style="Card.TFrame", padding=12)
        self.sales_chart_card.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=(0, 10))
        ttk.Label(self.sales_chart_card, text="Units Sold by Date", style="SectionTitle.TLabel").pack(anchor="w")

        self.sold_mix_chart_card = ttk.Frame(outer, style="Card.TFrame", padding=12)
        self.sold_mix_chart_card.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
        ttk.Label(self.sold_mix_chart_card, text="Sold Mix", style="SectionTitle.TLabel").pack(anchor="w")

        self.available_mix_chart_card = ttk.Frame(outer, style="Card.TFrame", padding=12)
        self.available_mix_chart_card.grid(row=1, column=1, sticky="nsew")
        ttk.Label(self.available_mix_chart_card, text="Available Mix", style="SectionTitle.TLabel").pack(anchor="w")

        self.sales_chart_canvas = None
        self.sold_mix_chart_canvas = None
        self.available_mix_chart_canvas = None

    def _create_tree(self, parent, columns: tuple[str, ...], headings: tuple[str, ...], height: int = 12):
        frame = ttk.Frame(parent, style="Card.TFrame")
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=height)
        vsb = ttk.Scrollbar(frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        right_aligned_columns = {
            "id",
            "value",
            "number_of_units_sold_in_sale",
            "sale_count",
            "units_sold",
            "face_value_total",
            "units_allocated",
        }
        for column, heading in zip(columns, headings):
            tree.heading(column, text=heading)
            width = 110
            if column in {"output_xlsx", "output_pdf", "sale_code", "serial_number", "pin_number"}:
                width = 180
            elif column in {"client", "dealer", "supplier", "invoice"}:
                width = 140
            anchor = "e" if column in right_aligned_columns else "center"
            tree.column(column, width=width, anchor=anchor)
        return frame

    def _configure_inventory_sorting(self, tree):
        for column in tree["columns"]:
            tree.heading(column, command=lambda col=column: self.sort_inventory_by(col))

    def _inventory_sort_key(self, row: sqlite3.Row, column: str):
        value = row[column]
        if value is None:
            return "" if column not in {"id", "value", "number_of_units_sold_in_sale"} else -1
        if column in {"id", "value", "number_of_units_sold_in_sale"}:
            return int(value)
        return str(value).strip().lower()

    def sort_inventory_by(self, column: str):
        if self.inventory_sort_column == column:
            self.inventory_sort_reverse = not self.inventory_sort_reverse
        else:
            self.inventory_sort_column = column
            self.inventory_sort_reverse = True if column == "date_of_sale" else False
        self.refresh_inventory()

    def pick_history_date(self):
        selected = ask_date(self, self.history_date_search_var.get().strip() or None)
        if selected:
            self.history_date_search_var.set(selected)

    def refresh_all(self):
        self.refresh_dashboard()
        self.refresh_sales()
        self.refresh_inventory()
        self.refresh_history()
        self.refresh_analytics()
        self.refresh_sale_available()
        self.update_stock_count_button_state()

    def refresh_dashboard(self):
        metrics = self.db.get_dashboard_metrics()
        self.stat_total_cards.set_value(str(metrics["total_cards"]))
        self.stat_available.set_value(str(metrics["available_cards"]))
        self.stat_sold.set_value(str(metrics["sold_cards"]))
        self.stat_sales_count.set_value(str(metrics["sales_count"]))

        tree = self.dashboard_sales_tree.children.get("!treeview")
        tree.delete(*tree.get_children())
        for row in self.db.get_recent_sales_summary():
            tree.insert("", END, values=(row["sale_date"], row["sale_count"], row["units_sold"], row["face_value_total"]))

        status_lines = [
            f"Available inventory face value: {metrics['available_value']}",
            f"Sold inventory face value: {metrics['sold_value']}",
            f"Total units sold through recorded sales: {metrics['units_sold_total']}",
        ]
        available_mix = self.db.get_value_breakdown(sold=False)
        if available_mix:
            status_lines.append("")
            status_lines.append("Available cards by value:")
            for row in available_mix:
                status_lines.append(f"- Value {row['value']}: {row['qty']} cards")

        self.quick_status_text.configure(state="normal")
        self.quick_status_text.delete("1.0", END)
        self.quick_status_text.insert("1.0", "\n".join(status_lines))
        self.quick_status_text.configure(state="disabled")

    def browse_import_file(self):
        path = filedialog.askopenfilename(
            title="Choose scratchcard Excel file",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            self.import_path_var.set(path)

    def import_excel(self):
        raw_path = self.import_path_var.get().strip()
        if not raw_path:
            messagebox.showerror(APP_TITLE, "Please choose an Excel file first.")
            return

        try:
            import_result = self.db.import_scratchcards(Path(raw_path))
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            self.import_status_var.set(f"Import failed: {exc}")
            return

        inserted = import_result["inserted"]
        updated = import_result["updated"]
        inserted_by_value = import_result["inserted_by_value"]
        updated_by_value = import_result["updated_by_value"]
        summary_lines = [f"Import complete. Inserted {inserted} row(s), updated {updated} row(s).", ""]
        summary_lines.append("Inserted by value:")
        for value in ALLOWED_VALUES:
            summary_lines.append(f"- {value}: {inserted_by_value.get(value, 0)}")
        summary_lines.append("")
        summary_lines.append("Updated by value:")
        for value in ALLOWED_VALUES:
            summary_lines.append(f"- {value}: {updated_by_value.get(value, 0)}")

        self.import_status_var.set(summary_lines[0])
        self.refresh_all()
        messagebox.showinfo(APP_TITLE, "\n".join(summary_lines))

    def refresh_sale_available(self):
        try:
            value = int(self.sale_value_var.get())
        except ValueError:
            self.sale_available_var.set("Available cards for selected value: 0")
            return
        available = self.db.available_count_for_value(value)
        self.sale_available_var.set(f"Available cards for selected value: {available}")

    def update_stock_count_button_state(self):
        rows = self.db.stock_count_summary()
        counts = {row["value"]: row["qty"] for row in rows}
        is_low_stock = any(counts.get(value, 0) < 10 for value in ALLOWED_VALUES)
        self.stock_count_button.configure(style="Alert.TButton" if is_low_stock else "Secondary.TButton")

    def show_stock_count(self):
        rows = self.db.stock_count_summary()
        counts = {row["value"]: row["qty"] for row in rows}
        lines = ["Rows with no sale date, client, or dealer:", ""]
        for value in ALLOWED_VALUES:
            qty = counts.get(value, 0)
            suffix = "  LOW STOCK" if qty < 10 else ""
            lines.append(f"Value {value}: {qty}{suffix}")
        messagebox.showinfo(APP_TITLE, "\n".join(lines))

    def run_sale(self):
        client = self.sale_client_var.get().strip()
        dealer = self.sale_dealer_var.get().strip()
        invoice = self.sale_invoice_var.get().strip()
        sale_date = self.sale_date_var.get().strip()
        units_text = self.sale_units_var.get().strip()
        value_text = self.sale_value_var.get().strip()

        if not client or not dealer or not invoice:
            messagebox.showerror(APP_TITLE, "Client name, dealer, and invoice are required.")
            return
        try:
            units = int(units_text)
            value = int(value_text)
        except ValueError:
            messagebox.showerror(APP_TITLE, "Value and number of units must be valid numbers.")
            return
        if units <= 0:
            messagebox.showerror(APP_TITLE, "Number of units must be greater than zero.")
            return

        try:
            sale_result = self.db.create_sale(client=client, dealer=dealer, invoice=invoice, sale_date=sale_date, value=value, units=units)
            self.reports.create_excel_report(sale_result, client=client, sale_date=normalize_date(sale_date), value=value)
            self.reports.create_pdf_report(sale_result, client=client, dealer=dealer, sale_date=normalize_date(sale_date), value=value)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            self.sale_status_var.set(f"Sale failed: {exc}")
            return

        self.sale_status_var.set(
            f"Sale {sale_result.sale_code} created for invoice {invoice}. Excel: {sale_result.excel_path.name} | PDF: {sale_result.pdf_path.name}"
        )
        self.refresh_all()
        messagebox.showinfo(
            APP_TITLE,
            "Sale completed successfully.\n\n"
            f"Excel: {sale_result.excel_path}\n"
            f"PDF: {sale_result.pdf_path}",
        )

    def refresh_sales(self):
        tree = self.sales_tree.children.get("!treeview")
        tree.delete(*tree.get_children())
        for row in self.db.list_sales()[:20]:
            tree.insert("", END, values=(row["sale_code"], row["sale_date"], row["client"], row["dealer"], row["invoice"], row["value"], row["units_allocated"]))

    def refresh_inventory(self):
        tree = self.inventory_tree.children.get("!treeview")
        tree.delete(*tree.get_children())
        rows = self.db.list_inventory(
            value_filter=self.inventory_value_var.get(),
            status_filter=self.inventory_status_var.get(),
            search_text=self.inventory_search_var.get(),
        )
        rows = sorted(
            rows,
            key=lambda row: self._inventory_sort_key(row, self.inventory_sort_column),
            reverse=self.inventory_sort_reverse,
        )
        self.inventory_results_var.set(f"Result of search: {len(rows)}")
        for row in rows:
            tree.insert(
                "",
                END,
                values=(
                    row["id"],
                    row["serial_number"],
                    row["pin_number"],
                    row["value"],
                    row["date_of_purchase"],
                    row["expiry_date"],
                    row["supplier"],
                    row["date_of_sale"],
                    row["number_of_units_sold_in_sale"],
                    row["client"],
                    row["dealer"],
                    row["invoice"],
                ),
            )

    def refresh_history(self):
        tree = self.history_tree.children.get("!treeview")
        tree.delete(*tree.get_children())
        rows = self.db.list_history_groups(
            client_search=self.history_client_search_var.get(),
            date_search=self.history_date_search_var.get(),
            dealer_search=self.history_dealer_search_var.get(),
            invoice_search=self.history_invoice_search_var.get(),
        )
        self.selected_history_group = None
        self.history_detail_var.set("Click a history row to see the sale details.")
        detail_tree = self.history_detail_tree.children.get("!treeview")
        detail_tree.delete(*detail_tree.get_children())
        for row in rows:
            values_sold = ", ".join(sorted((row["value_mix"] or "").split(","), key=lambda item: int(item))) if row["value_mix"] else ""
            tree.insert(
                "",
                END,
                values=(
                    row["sale_date"],
                    row["client"],
                    row["dealer"],
                    row["invoice"],
                    row["units_sold"],
                    row["face_value_total"],
                    values_sold,
                ),
            )

    def on_history_select(self, _event=None):
        tree = self.history_tree.children.get("!treeview")
        selection = tree.selection()
        if not selection:
            self.selected_history_group = None
            return
        sale_date, client, dealer, invoice, units_sold, _face_value_total, values_sold = tree.item(selection[0], "values")
        self.selected_history_group = (sale_date, client, dealer, invoice)
        detail_rows = self.db.get_history_group_rows(sale_date, client, dealer, invoice)
        self.history_detail_var.set(f"{sale_date} | {client} | {dealer} | Invoice: {invoice} | {units_sold} unit(s) | Values: {values_sold}")
        detail_tree = self.history_detail_tree.children.get("!treeview")
        detail_tree.delete(*detail_tree.get_children())
        for row in detail_rows:
            detail_tree.insert(
                "",
                END,
                values=(
                    row["serial_number"],
                    row["pin_number"],
                    row["value"],
                    row["invoice"],
                    row["supplier"],
                    row["date_of_purchase"],
                    row["expiry_date"],
                ),
            )

    def export_selected_history_sale(self):
        if self.selected_history_group is None:
            messagebox.showwarning(APP_TITLE, "Please click a history row first.")
            return
        sale_date, client, dealer, invoice = self.selected_history_group
        rows = self.db.get_history_group_rows(sale_date, client, dealer, invoice)
        try:
            excel_path, pdf_path = self.reports.export_group_rows(rows, client=client, dealer=dealer, sale_date=sale_date)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            return
        messagebox.showinfo(APP_TITLE, f"Sale detail files exported.\n\nExcel: {excel_path}\nPDF: {pdf_path}")

    def refresh_analytics(self):
        self._draw_sales_chart()
        self._draw_mix_chart(sold=True, title="Sold Mix", master=self.sold_mix_chart_card, canvas_attr="sold_mix_chart_canvas")
        self._draw_mix_chart(sold=False, title="Available Mix", master=self.available_mix_chart_card, canvas_attr="available_mix_chart_canvas")

    def _draw_sales_chart(self):
        summary = self.db.get_monthly_sales_summary()
        labels = [row["sale_month"] for row in summary]
        values = [row["units_sold"] for row in summary]
        fig = Figure(figsize=(8.6, 4.2), dpi=100, facecolor="white")
        ax = fig.add_subplot(111)
        if values:
            x_positions = list(range(len(labels)))
            ax.plot(
                x_positions,
                values,
                color=GREEN,
                linewidth=2.2,
                marker="o",
                markersize=4.5,
                markerfacecolor=NAVY,
                markeredgecolor=NAVY,
            )
            ax.set_xticks(x_positions)
            ax.set_xticklabels(labels)
            ax.set_ylim(0, max(values) + max(1, int(max(values) * 0.15)))
        else:
            ax.text(0.5, 0.5, "No sold inventory yet", ha="center", va="center", fontsize=12, color=MUTED, transform=ax.transAxes)
            ax.set_xticks([])
            ax.set_yticks([])
        ax.set_title("Units Sold by Month", color=TEXT_DARK)
        ax.tick_params(axis="x", rotation=90, labelsize=7)
        ax.tick_params(axis="y", labelsize=8)
        ax.grid(axis="y", color="#d9e7ef", linewidth=0.8)
        ax.set_facecolor("white")
        for spine in ax.spines.values():
            spine.set_color("#c3d6e2")
        fig.tight_layout()

        if self.sales_chart_canvas is not None:
            self.sales_chart_canvas.get_tk_widget().destroy()
        self.sales_chart_canvas = FigureCanvasTkAgg(fig, master=self.sales_chart_card)
        self.sales_chart_canvas.draw()
        self.sales_chart_canvas.get_tk_widget().pack(fill=BOTH, expand=True, pady=(10, 0))

    def _draw_mix_chart(self, sold: bool, title: str, master, canvas_attr: str):
        breakdown = self.db.get_value_breakdown(sold=sold)
        fig = Figure(figsize=(4.6, 3.8), dpi=100, facecolor="white")
        ax = fig.add_subplot(111)
        labels = [f"{row['value']} ({row['qty']})" for row in breakdown if (row["qty"] or 0) > 0]
        values = [row["qty"] for row in breakdown if (row["qty"] or 0) > 0]
        if values:
            ax.pie(
                values,
                labels=labels,
                colors=[NAVY, GREEN, "#1ea7d7", "#44c29e", "#90d2bf"],
                startangle=110,
            )
        else:
            ax.text(0.5, 0.5, f"No {'sold' if sold else 'available'} inventory yet", ha="center", va="center", fontsize=12, color=MUTED, transform=ax.transAxes)
            ax.set_xticks([])
            ax.set_yticks([])
            for spine in ax.spines.values():
                spine.set_visible(False)
        ax.set_title(title, color=TEXT_DARK)
        fig.tight_layout()

        current_canvas = getattr(self, canvas_attr)
        if current_canvas is not None:
            current_canvas.get_tk_widget().destroy()
        new_canvas = FigureCanvasTkAgg(fig, master=master)
        new_canvas.draw()
        new_canvas.get_tk_widget().pack(fill=BOTH, expand=True, pady=(10, 0))
        setattr(self, canvas_attr, new_canvas)

    def on_inventory_select(self, _event=None):
        tree = self.inventory_tree.children.get("!treeview")
        selection = tree.selection()
        if not selection:
            self.selected_inventory_row_id = None
            return
        values = tree.item(selection[0], "values")
        self.selected_inventory_row_id = int(values[0])

    def edit_selected_inventory(self):
        if self.selected_inventory_row_id is None:
            messagebox.showwarning(APP_TITLE, "Please select an inventory row first.")
            return

        field = simpledialog.askstring(
            APP_TITLE,
            "Enter the field to edit:\n"
            "date_of_purchase\nexpiry_date\nsupplier\ndate_of_sale\nnumber_of_units_sold_in_sale\nclient\ndealer\ninvoice",
            parent=self,
        )
        if not field:
            return
        value = simpledialog.askstring(APP_TITLE, f"Enter new value for {field}:", parent=self)
        try:
            self.db.update_field(self.selected_inventory_row_id, field.strip(), value)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            return
        self.refresh_all()
        messagebox.showinfo(APP_TITLE, "Record updated successfully.")


if __name__ == "__main__":
    app = ThurayaApp()
    app.mainloop()













