import calendar
import sqlite3
from datetime import date, datetime
from pathlib import Path
import tkinter as tk
from tkinter import BOTH, END, VERTICAL, filedialog, messagebox, ttk

NAVY = "#0c2d48"
OFF_WHITE = "#f4f8fb"
CARD_BG = "#ffffff"
TEXT_DARK = "#183247"
MUTED = "#607587"
NEGATIVE = "#c0392b"
POSITIVE = "#1f5f99"

RCS_INVOICE_PARTIES = ["Xtra-Link", "Marlink", "RCSi", "Other"]
RCS_PAYMENT_PARTIES = ["DFCU", "Other"]
MONTHLY_CLIENTS = [
    "Embassy of Sweden",
    "Totalenergies EP",
    "CFAO Mob",
    "Totalenergies Marketing",
    "Japanese Embassy",
    "Invesco",
]


def normalize_date(value) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def format_amount(value) -> str:
    if value in (None, ""):
        return ""
    number = parse_number(value, 0.0)
    return f"{number:,.2f}" if number % 1 else f"{int(number):,}"


def format_usd(value) -> str:
    if value in (None, ""):
        return ""
    number = parse_number(value, 0.0) or 0.0
    return f"${number:,.2f}"


def format_ugx(value) -> str:
    if value in (None, ""):
        return ""
    return f"/={format_amount(value)}"


def execute_checked(conn: sqlite3.Connection, query: str, params: tuple, missing_message: str):
    cursor = conn.execute(query, params)
    if cursor.rowcount == 0:
        raise ValueError(missing_message)
    return cursor


def parse_number(value, default=None):
    if value in (None, ""):
        return default
    cleaned = str(value).replace(",", "").replace("$", "").replace("/=", "").strip()
    if cleaned.lower() in {"none", "null", "nil"}:
        return default
    return float(cleaned) if cleaned else default


def strip_excel_text_prefix(value) -> str:
    text = str(value or "").strip()
    if text.startswith("'"):
        text = text[1:]
    return text.strip()


def normalize_excel_header(value) -> str:
    text = strip_excel_text_prefix(value).replace("\xa0", " ").replace("_", " ").replace("-", " ")
    return " ".join(text.lower().split())


def clean_excel_text(value) -> str:
    return strip_excel_text_prefix(value).replace("\xa0", " ").strip()


class CalendarPicker(tk.Toplevel):
    def __init__(self, parent, initial_date: str | None = None):
        super().__init__(parent)
        self.title("Choose Date")
        self.resizable(False, False)
        self.transient(parent)
        self.result: str | None = None
        parsed = normalize_date(initial_date) or date.today().isoformat()
        current = datetime.strptime(parsed, "%Y-%m-%d").date()
        self.displayed_year = current.year
        self.displayed_month = current.month
        self._build_ui()
        self._render_calendar()
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build_ui(self):
        header = ttk.Frame(self, padding=10)
        header.pack(fill="x")
        ttk.Button(header, text="<", width=3, command=self._prev_month).pack(side="left")
        self.month_label = ttk.Label(header, font=("Segoe UI Semibold", 11))
        self.month_label.pack(side="left", expand=True)
        ttk.Button(header, text=">", width=3, command=self._next_month).pack(side="right")

        days = ttk.Frame(self, padding=(10, 0, 10, 10))
        days.pack(fill="both", expand=True)
        for idx, name in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
            ttk.Label(days, text=name, anchor="center").grid(row=0, column=idx, padx=2, pady=(0, 6), sticky="nsew")
            days.columnconfigure(idx, weight=1)
        self.days_frame = days

    def _render_calendar(self):
        for child in list(self.days_frame.grid_slaves()):
            if int(child.grid_info().get("row", 0)) > 0:
                child.destroy()
        self.month_label.configure(text=f"{calendar.month_name[self.displayed_month]} {self.displayed_year}")
        for row_idx, week in enumerate(calendar.monthcalendar(self.displayed_year, self.displayed_month), start=1):
            for col_idx, day_num in enumerate(week):
                if day_num == 0:
                    ttk.Label(self.days_frame, text="").grid(row=row_idx, column=col_idx, padx=2, pady=2, sticky="nsew")
                    continue
                ttk.Button(self.days_frame, text=str(day_num), width=4, command=lambda d=day_num: self._select_day(d)).grid(
                    row=row_idx, column=col_idx, padx=2, pady=2, sticky="nsew"
                )

    def _prev_month(self):
        if self.displayed_month == 1:
            self.displayed_year -= 1
            self.displayed_month = 12
        else:
            self.displayed_month -= 1
        self._render_calendar()

    def _next_month(self):
        if self.displayed_month == 12:
            self.displayed_year += 1
            self.displayed_month = 1
        else:
            self.displayed_month += 1
        self._render_calendar()

    def _select_day(self, day_num: int):
        self.result = date(self.displayed_year, self.displayed_month, day_num).isoformat()
        self.destroy()


def ask_date(parent, initial_date: str | None = None) -> str | None:
    picker = CalendarPicker(parent, initial_date=initial_date)
    parent.wait_window(picker)
    return picker.result


class RcsDataService:
    def __init__(self, db_path: Path):
        self.db_path = Path(db_path)
        self._ensure_schema()

    def connect(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _ensure_schema(self):
        with self.connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS rcs_invoices (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    invoice_date TEXT NOT NULL,
                    reference TEXT NOT NULL,
                    value REAL NOT NULL,
                    party TEXT NOT NULL,
                    comment TEXT,
                    amount_paid REAL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS rcs_payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    payment_date TEXT NOT NULL,
                    reference TEXT NOT NULL,
                    value REAL NOT NULL,
                    party TEXT NOT NULL,
                    comment TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS rcs_monthly_invoices (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    invoice_date TEXT NOT NULL,
                    client TEXT NOT NULL,
                    value_ugx REAL NOT NULL,
                    exchange_rate REAL NOT NULL DEFAULT 3900,
                    value_usd REAL NOT NULL,
                    invoice_number TEXT NOT NULL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS marlink (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    entry_date TEXT NOT NULL,
                    iridium_number TEXT NOT NULL,
                    voucher_type TEXT NOT NULL,
                    client TEXT NOT NULL,
                    cost REAL NOT NULL,
                    marlink_invoice TEXT,
                    marlink_invoice_value REAL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS iridium_vouchers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    units INTEGER NOT NULL,
                    cost REAL NOT NULL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS iridium_sim (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    msisdn TEXT NOT NULL,
                    iccid TEXT NOT NULL,
                    client TEXT NOT NULL,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                """
            )
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_iridium_vouchers_name ON iridium_vouchers(name)")
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_iridium_sim_msisdn ON iridium_sim(msisdn)")
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_iridium_sim_iccid ON iridium_sim(iccid)")
            conn.commit()

    def list_rcs_invoices(self, date_search: str = "", reference_search: str = "", party: str = "All"):
        query = "SELECT * FROM rcs_invoices WHERE 1=1"
        params = []
        if date_search.strip():
            query += " AND invoice_date LIKE ?"
            params.append(f"%{date_search.strip()}%")
        if reference_search.strip():
            query += " AND reference LIKE ?"
            params.append(f"%{reference_search.strip()}%")
        if party and party != "All":
            query += " AND party = ?"
            params.append(party)
        query += " ORDER BY invoice_date DESC, id DESC"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def create_rcs_invoice(self, payload: dict):
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO rcs_invoices (invoice_date, reference, value, party, comment, amount_paid, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    normalize_date(payload["invoice_date"]),
                    payload["reference"].strip(),
                    parse_number(payload["value"], 0.0),
                    payload["party"],
                    payload.get("comment") or None,
                    parse_number(payload.get("amount_paid"), None),
                    now_iso(),
                ),
            )
            conn.commit()

    def update_rcs_invoice(self, record_id: int, payload: dict):
        with self.connect() as conn:
            execute_checked(
                conn,
                """
                UPDATE rcs_invoices
                SET invoice_date = ?, reference = ?, value = ?, party = ?, comment = ?, amount_paid = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    normalize_date(payload["invoice_date"]),
                    payload["reference"].strip(),
                    parse_number(payload["value"], 0.0),
                    payload["party"],
                    payload.get("comment") or None,
                    parse_number(payload.get("amount_paid"), None),
                    now_iso(),
                    record_id,
                ),
                f"RCS invoice row {record_id} was not found.",
            )
            conn.commit()

    def delete_rcs_invoice(self, record_id: int):
        with self.connect() as conn:
            conn.execute("DELETE FROM rcs_invoices WHERE id = ?", (record_id,))
            conn.commit()

    def list_rcs_payments(self, date_search: str = "", reference_search: str = "", party: str = "All"):
        query = "SELECT * FROM rcs_payments WHERE 1=1"
        params = []
        if date_search.strip():
            query += " AND payment_date LIKE ?"
            params.append(f"%{date_search.strip()}%")
        if reference_search.strip():
            query += " AND reference LIKE ?"
            params.append(f"%{reference_search.strip()}%")
        if party and party != "All":
            query += " AND party = ?"
            params.append(party)
        query += " ORDER BY payment_date DESC, id DESC"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def create_rcs_payment(self, payload: dict):
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO rcs_payments (payment_date, reference, value, party, comment, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    normalize_date(payload["payment_date"]),
                    payload["reference"].strip(),
                    parse_number(payload["value"], 0.0),
                    payload["party"],
                    payload.get("comment") or None,
                    now_iso(),
                ),
            )
            conn.commit()

    def update_rcs_payment(self, record_id: int, payload: dict):
        with self.connect() as conn:
            execute_checked(
                conn,
                """
                UPDATE rcs_payments
                SET payment_date = ?, reference = ?, value = ?, party = ?, comment = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    normalize_date(payload["payment_date"]),
                    payload["reference"].strip(),
                    parse_number(payload["value"], 0.0),
                    payload["party"],
                    payload.get("comment") or None,
                    now_iso(),
                    record_id,
                ),
                f"RCS payment row {record_id} was not found.",
            )
            conn.commit()

    def delete_rcs_payment(self, record_id: int):
        with self.connect() as conn:
            conn.execute("DELETE FROM rcs_payments WHERE id = ?", (record_id,))
            conn.commit()

    def list_monthly_invoices(self, date_search: str = "", client: str = "All", invoice_number: str = ""):
        query = "SELECT * FROM rcs_monthly_invoices WHERE 1=1"
        params = []
        if date_search.strip():
            query += " AND invoice_date LIKE ?"
            params.append(f"%{date_search.strip()}%")
        if client and client != "All":
            query += " AND client = ?"
            params.append(client)
        if invoice_number.strip():
            query += " AND invoice_number LIKE ?"
            params.append(f"%{invoice_number.strip()}%")
        query += " ORDER BY invoice_date DESC, id DESC"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def create_monthly_invoice(self, payload: dict):
        client = payload["client"]
        exchange_rate = parse_number(payload.get("exchange_rate"), 3900.0)
        if client == "Totalenergies EP":
            value_usd = parse_number(payload.get("value_usd"), 0.0)
            value_ugx = value_usd * exchange_rate if exchange_rate else 0.0
        else:
            value_ugx = parse_number(payload["value_ugx"], 0.0)
            value_usd = value_ugx / exchange_rate if exchange_rate else 0.0
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO rcs_monthly_invoices (invoice_date, client, value_ugx, exchange_rate, value_usd, invoice_number, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    normalize_date(payload["invoice_date"]),
                    client,
                    value_ugx,
                    exchange_rate,
                    value_usd,
                    payload["invoice_number"].strip(),
                    now_iso(),
                ),
            )
            conn.commit()

    def update_monthly_invoice(self, record_id: int, payload: dict):
        client = payload["client"]
        exchange_rate = parse_number(payload.get("exchange_rate"), 3900.0)
        if client == "Totalenergies EP":
            value_usd = parse_number(payload.get("value_usd"), 0.0)
            value_ugx = value_usd * exchange_rate if exchange_rate else 0.0
        else:
            value_ugx = parse_number(payload["value_ugx"], 0.0)
            value_usd = value_ugx / exchange_rate if exchange_rate else 0.0
        with self.connect() as conn:
            execute_checked(
                conn,
                """
                UPDATE rcs_monthly_invoices
                SET invoice_date = ?, client = ?, value_ugx = ?, exchange_rate = ?, value_usd = ?, invoice_number = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    normalize_date(payload["invoice_date"]),
                    client,
                    value_ugx,
                    exchange_rate,
                    value_usd,
                    payload["invoice_number"].strip(),
                    now_iso(),
                    record_id,
                ),
                f"RCS monthly invoice row {record_id} was not found.",
            )
            conn.commit()

    def delete_monthly_invoice(self, record_id: int):
        with self.connect() as conn:
            conn.execute("DELETE FROM rcs_monthly_invoices WHERE id = ?", (record_id,))
            conn.commit()

    def list_marlink(self, date_search: str = "", iridium_number: str = "", voucher_type: str = "", client: str = "", marlink_invoice: str = ""):
        query = "SELECT * FROM marlink WHERE 1=1"
        params = []
        if date_search.strip():
            query += " AND entry_date LIKE ?"
            params.append(f"%{date_search.strip()}%")
        if iridium_number.strip():
            query += " AND iridium_number LIKE ?"
            params.append(f"%{iridium_number.strip()}%")
        if voucher_type.strip():
            query += " AND voucher_type LIKE ?"
            params.append(f"%{voucher_type.strip()}%")
        if client.strip():
            query += " AND client LIKE ?"
            params.append(f"%{client.strip()}%")
        if marlink_invoice.strip():
            query += " AND marlink_invoice LIKE ?"
            params.append(f"%{marlink_invoice.strip()}%")
        query += " ORDER BY entry_date DESC, id DESC"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def create_marlink(self, payload: dict):
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO marlink (entry_date, iridium_number, voucher_type, client, cost, marlink_invoice, marlink_invoice_value, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    normalize_date(payload["entry_date"]),
                    payload["iridium_number"].strip(),
                    payload["voucher_type"].strip(),
                    payload["client"].strip(),
                    parse_number(payload["cost"], 0.0),
                    (payload.get("marlink_invoice") or "").strip() or None,
                    parse_number(payload.get("marlink_invoice_value"), None),
                    now_iso(),
                ),
            )
            conn.commit()

    def import_marlink_excel(self, file_path: str) -> dict:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to import Marlink rows.") from exc

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Import file not found: {path}")

        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("The Excel file is empty.")

        header_map = {normalize_excel_header(value): idx for idx, value in enumerate(header_row) if value is not None}
        aliases = {
            "date": ["date", "entry date"],
            "iridium number": ["iridium number", "iridium no", "msisdn"],
            "voucher type": ["voucher type", "voucher", "voucher name"],
            "marlink invoice": ["marlink invoice", "invoice", "invoice number"],
        }
        resolved_headers: dict[str, int] = {}
        missing: list[str] = []
        for required_name, candidates in aliases.items():
            matched_idx = next((header_map[name] for name in candidates if name in header_map), None)
            if matched_idx is None:
                missing.append(required_name)
            else:
                resolved_headers[required_name] = matched_idx
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        inserted = 0
        updated = 0
        last_seen_date = ""
        last_seen_invoice = ""
        with self.connect() as conn:
            for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                raw_date = normalize_date(values[resolved_headers["date"]])
                entry_date = raw_date or last_seen_date
                iridium_number = clean_excel_text(values[resolved_headers["iridium number"]])
                voucher_type = clean_excel_text(values[resolved_headers["voucher type"]])
                raw_invoice = clean_excel_text(values[resolved_headers["marlink invoice"]])
                marlink_invoice = raw_invoice or last_seen_invoice

                if raw_date:
                    last_seen_date = raw_date
                if raw_invoice:
                    last_seen_invoice = raw_invoice

                if not entry_date and not iridium_number and not voucher_type and not marlink_invoice:
                    continue
                if not entry_date or not iridium_number or not voucher_type or not marlink_invoice:
                    missing_fields = []
                    if not entry_date:
                        missing_fields.append("Date")
                    if not iridium_number:
                        missing_fields.append("Iridium Number")
                    if not voucher_type:
                        missing_fields.append("Voucher Type")
                    if not marlink_invoice:
                        missing_fields.append("Marlink Invoice")
                    raise ValueError(f"Row {row_number}: missing {', '.join(missing_fields)}.")

                sim = conn.execute("SELECT client FROM iridium_sim WHERE msisdn = ?", (iridium_number,)).fetchone()
                if not sim:
                    placeholder_iccid = f"***{iridium_number}"
                    conn.execute(
                        """
                        INSERT INTO iridium_sim (msisdn, iccid, client, updated_at)
                        VALUES (?, ?, ?, ?)
                        """,
                        (iridium_number, placeholder_iccid, "UNKNOWN CLIENT", now_iso()),
                    )
                    client = "UNKNOWN CLIENT"
                else:
                    client = sim["client"]

                voucher = conn.execute("SELECT cost FROM iridium_vouchers WHERE name = ?", (voucher_type,)).fetchone()
                if not voucher:
                    raise ValueError(f"Row {row_number}: Voucher Type not found in parameters: {voucher_type}")

                cost = parse_number(voucher["cost"], 0.0) or 0.0

                existing = conn.execute(
                    """
                    SELECT id FROM marlink
                    WHERE entry_date = ? AND iridium_number = ? AND voucher_type = ? AND marlink_invoice = ?
                    """,
                    (entry_date, iridium_number, voucher_type, marlink_invoice),
                ).fetchone()
                if existing:
                    conn.execute(
                        """
                        UPDATE marlink
                        SET client = ?, cost = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (client, cost, now_iso(), existing["id"]),
                    )
                    updated += 1
                else:
                    conn.execute(
                        """
                        INSERT INTO marlink (entry_date, iridium_number, voucher_type, client, cost, marlink_invoice, marlink_invoice_value, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (entry_date, iridium_number, voucher_type, client, cost, marlink_invoice, None, now_iso()),
                    )
                    inserted += 1
            conn.commit()
        return {"inserted": inserted, "updated": updated}

    def update_marlink(self, record_id: int, payload: dict):
        with self.connect() as conn:
            execute_checked(
                conn,
                """
                UPDATE marlink
                SET entry_date = ?, iridium_number = ?, voucher_type = ?, client = ?, cost = ?, marlink_invoice = ?, marlink_invoice_value = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    normalize_date(payload["entry_date"]),
                    payload["iridium_number"].strip(),
                    payload["voucher_type"].strip(),
                    payload["client"].strip(),
                    parse_number(payload["cost"], 0.0),
                    (payload.get("marlink_invoice") or "").strip() or None,
                    parse_number(payload.get("marlink_invoice_value"), None),
                    now_iso(),
                    record_id,
                ),
                f"Marlink row {record_id} was not found.",
            )
            conn.commit()

    def delete_marlink(self, record_id: int):
        with self.connect() as conn:
            conn.execute("DELETE FROM marlink WHERE id = ?", (record_id,))
            conn.commit()

    def list_iridium_vouchers(self, name: str = ""):
        query = "SELECT * FROM iridium_vouchers WHERE 1=1"
        params = []
        if name.strip():
            query += " AND name LIKE ?"
            params.append(f"%{name.strip()}%")
        query += " ORDER BY name COLLATE NOCASE ASC, id DESC"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def create_iridium_voucher(self, payload: dict):
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO iridium_vouchers (name, units, cost, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                (
                    payload["name"].strip(),
                    int(parse_number(payload["units"], 0) or 0),
                    parse_number(payload["cost"], 0.0),
                    now_iso(),
                ),
            )
            conn.commit()

    def update_iridium_voucher(self, record_id: int, payload: dict):
        with self.connect() as conn:
            execute_checked(
                conn,
                """
                UPDATE iridium_vouchers
                SET name = ?, units = ?, cost = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    payload["name"].strip(),
                    int(parse_number(payload["units"], 0) or 0),
                    parse_number(payload["cost"], 0.0),
                    now_iso(),
                    record_id,
                ),
                f"Iridium voucher row {record_id} was not found.",
            )
            conn.commit()

    def delete_iridium_voucher(self, record_id: int):
        with self.connect() as conn:
            conn.execute("DELETE FROM iridium_vouchers WHERE id = ?", (record_id,))
            conn.commit()

    def voucher_names(self) -> list[str]:
        with self.connect() as conn:
            return [row["name"] for row in conn.execute("SELECT name FROM iridium_vouchers ORDER BY name COLLATE NOCASE ASC").fetchall()]

    def voucher_by_name(self, name: str) -> sqlite3.Row | None:
        with self.connect() as conn:
            return conn.execute("SELECT * FROM iridium_vouchers WHERE name = ?", (name.strip(),)).fetchone()

    def list_iridium_sims(self, msisdn: str = "", iccid: str = "", client: str = ""):
        query = "SELECT * FROM iridium_sim WHERE 1=1"
        params = []
        if msisdn.strip():
            query += " AND msisdn LIKE ?"
            params.append(f"%{msisdn.strip()}%")
        if iccid.strip():
            query += " AND iccid LIKE ?"
            params.append(f"%{iccid.strip()}%")
        if client.strip():
            query += " AND client LIKE ?"
            params.append(f"%{client.strip()}%")
        query += " ORDER BY msisdn COLLATE NOCASE ASC, id DESC"
        with self.connect() as conn:
            return conn.execute(query, params).fetchall()

    def create_iridium_sim(self, payload: dict):
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO iridium_sim (msisdn, iccid, client, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                (
                    payload["msisdn"].strip(),
                    payload["iccid"].strip(),
                    payload["client"].strip(),
                    now_iso(),
                ),
            )
            conn.commit()

    def update_iridium_sim(self, record_id: int, payload: dict):
        with self.connect() as conn:
            execute_checked(
                conn,
                """
                UPDATE iridium_sim
                SET msisdn = ?, iccid = ?, client = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    payload["msisdn"].strip(),
                    payload["iccid"].strip(),
                    payload["client"].strip(),
                    now_iso(),
                    record_id,
                ),
                f"Iridium SIM row {record_id} was not found.",
            )
            conn.commit()

    def delete_iridium_sim(self, record_id: int):
        with self.connect() as conn:
            conn.execute("DELETE FROM iridium_sim WHERE id = ?", (record_id,))
            conn.commit()

    def sim_msisdns(self) -> list[str]:
        with self.connect() as conn:
            return [row["msisdn"] for row in conn.execute("SELECT msisdn FROM iridium_sim ORDER BY msisdn COLLATE NOCASE ASC").fetchall()]

    def sim_by_msisdn(self, msisdn: str) -> sqlite3.Row | None:
        with self.connect() as conn:
            return conn.execute("SELECT * FROM iridium_sim WHERE msisdn = ?", (msisdn.strip(),)).fetchone()

    def match_sim_msisdn(self, typed_value: str) -> sqlite3.Row | None:
        text = typed_value.strip()
        if not text:
            return None
        with self.connect() as conn:
            exact = conn.execute("SELECT * FROM iridium_sim WHERE msisdn = ?", (text,)).fetchone()
            if exact:
                return exact
            matches = conn.execute(
                "SELECT * FROM iridium_sim WHERE msisdn LIKE ? ORDER BY msisdn COLLATE NOCASE ASC",
                (f"{text}%",),
            ).fetchall()
            if len(matches) == 1:
                return matches[0]
        return None

    def import_iridium_sims_excel(self, file_path: str) -> dict:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to import Iridium SIMs.") from exc

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Import file not found: {path}")

        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("The Excel file is empty.")

        header_map = {str(value).strip().lower(): idx for idx, value in enumerate(header_row) if value is not None}
        required = {"msisdn", "iccid", "client"}
        missing = [name for name in required if name not in header_map]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        inserted = 0
        updated = 0
        with self.connect() as conn:
            for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                msisdn = strip_excel_text_prefix(values[header_map["msisdn"]])
                iccid = strip_excel_text_prefix(values[header_map["iccid"]])
                client = str(values[header_map["client"]] or "").strip()
                if not msisdn and not iccid and not client:
                    continue
                if not msisdn or not iccid or not client:
                    raise ValueError(f"Row {row_number}: MSISDN, ICCID, and Client are required.")

                existing = conn.execute(
                    "SELECT id FROM iridium_sim WHERE msisdn = ? OR iccid = ?",
                    (msisdn, iccid),
                ).fetchone()
                if existing:
                    conn.execute(
                        """
                        UPDATE iridium_sim
                        SET msisdn = ?, iccid = ?, client = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (msisdn, iccid, client, now_iso(), existing["id"]),
                    )
                    updated += 1
                else:
                    conn.execute(
                        """
                        INSERT INTO iridium_sim (msisdn, iccid, client, updated_at)
                        VALUES (?, ?, ?, ?)
                        """,
                        (msisdn, iccid, client, now_iso()),
                    )
                    inserted += 1
            conn.commit()
        return {"inserted": inserted, "updated": updated}

    def build_supplier_ledger(self, invoice_party: str):
        with self.connect() as conn:
            invoices = conn.execute(
                "SELECT invoice_date, party, reference, value FROM rcs_invoices WHERE party = ? ORDER BY invoice_date ASC, id ASC",
                (invoice_party,),
            ).fetchall()
            payments = conn.execute(
                "SELECT payment_date, reference, value FROM rcs_payments ORDER BY payment_date ASC, id ASC"
            ).fetchall()

        events = []
        for row in invoices:
            events.append(
                {
                    "sort_date": row["invoice_date"] or "",
                    "invoice_date": row["invoice_date"] or "",
                    "supplier": row["party"],
                    "reference": row["reference"],
                    "invoice_value": parse_number(row["value"], 0.0) or 0.0,
                    "payment_date": "",
                    "payment_value": 0.0,
                    "delta": parse_number(row["value"], 0.0) or 0.0,
                }
            )
        for row in payments:
            events.append(
                {
                    "sort_date": row["payment_date"] or "",
                    "invoice_date": "",
                    "supplier": "",
                    "reference": row["reference"],
                    "invoice_value": 0.0,
                    "payment_date": row["payment_date"] or "",
                    "payment_value": parse_number(row["value"], 0.0) or 0.0,
                    "delta": -(parse_number(row["value"], 0.0) or 0.0),
                }
            )

        events.sort(key=lambda item: (item["sort_date"], item["invoice_date"] == "", item["reference"]))
        balance = 0.0
        for item in events:
            balance += item["delta"]
            item["balance"] = balance
        events.reverse()

        total_invoices = sum(item["invoice_value"] for item in events)
        total_payments = sum(item["payment_value"] for item in events)
        return {
            "rows": events,
            "total_invoices": total_invoices,
            "total_payments": total_payments,
            "balance": total_invoices - total_payments,
        }


class RecordEditorDialog(tk.Toplevel):
    def __init__(self, parent, title: str, form_fields: list[dict], initial: dict | None = None, on_back=None):
        super().__init__(parent)
        self.title(title)
        self.transient(parent)
        self.resizable(False, False)
        self.result = None
        self.form_fields = form_fields
        self.initial = initial or {}
        self.on_back = on_back
        self.vars: dict[str, tk.StringVar] = {}
        self.widgets: dict[str, ttk.Entry | ttk.Combobox] = {}
        self._build_ui()
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build_ui(self):
        body = ttk.Frame(self, padding=16)
        body.pack(fill="both", expand=True)
        row = 0
        for field in self.form_fields:
            key = field["key"]
            var = tk.StringVar(value=str(self.initial.get(key, field.get("default", "")) or ""))
            self.vars[key] = var
            ttk.Label(body, text=field["label"]).grid(row=row, column=0, sticky="w", pady=(8 if row else 0, 4))
            if field.get("type") == "dropdown":
                widget = ttk.Combobox(body, values=field["options"], textvariable=var, state="readonly", width=30)
                widget.grid(row=row + 1, column=0, sticky="ew")
            else:
                widget = ttk.Entry(body, textvariable=var, width=34, state="readonly" if field.get("readonly") else "normal")
                widget.grid(row=row + 1, column=0, sticky="ew")
                if field.get("type") == "date":
                    ttk.Button(body, text="Calendar", command=lambda v=var: self._pick_date(v)).grid(row=row + 1, column=1, padx=(8, 0))
            self.widgets[key] = widget
            row += 2

        if "value_usd" in self.vars and "value_ugx" in self.vars and "exchange_rate" in self.vars:
            self.vars["value_ugx"].trace_add("write", lambda *_: self._update_usd())
            self.vars["exchange_rate"].trace_add("write", lambda *_: self._update_usd())
            if "client" in self.vars:
                self.vars["client"].trace_add("write", lambda *_: self._toggle_monthly_fields())
                self._toggle_monthly_fields()
            self._update_usd()

        actions = ttk.Frame(body)
        actions.grid(row=row, column=0, columnspan=2, sticky="e", pady=(14, 0))
        if self.on_back is not None:
            ttk.Button(actions, text="Back", command=self._go_back).pack(side="left", padx=(0, 8))
        ttk.Button(actions, text="Cancel", command=self.destroy).pack(side="right")
        ttk.Button(actions, text="Save", command=self._save).pack(side="right", padx=(0, 8))
        body.columnconfigure(0, weight=1)

    def _pick_date(self, var: tk.StringVar):
        selected = ask_date(self, var.get().strip() or None)
        if selected:
            var.set(selected)

    def _update_usd(self):
        try:
            client = self.vars["client"].get().strip() if "client" in self.vars else ""
            rate = parse_number(self.vars["exchange_rate"].get(), 0.0)
            if client == "Totalenergies EP":
                usd = parse_number(self.vars["value_usd"].get(), 0.0)
                ugx = usd * rate if rate else 0
                self.vars["value_ugx"].set(f"{ugx:.2f}")
            else:
                ugx = parse_number(self.vars["value_ugx"].get(), 0.0)
                usd = ugx / rate if rate else 0
        except ValueError:
            usd = 0
        self.vars["value_usd"].set(f"{usd:.2f}")

    def _go_back(self):
        self.destroy()
        self.on_back()

    def _toggle_monthly_fields(self):
        if "client" not in self.vars:
            return
        client = self.vars["client"].get().strip()
        ugx_widget = self.widgets.get("value_ugx")
        usd_widget = self.widgets.get("value_usd")
        if ugx_widget is None or usd_widget is None:
            return
        if client == "Totalenergies EP":
            ugx_widget.configure(state="readonly")
            usd_widget.configure(state="normal")
        else:
            ugx_widget.configure(state="normal")
            usd_widget.configure(state="readonly")

    def _save(self):
        payload = {}
        for field in self.form_fields:
            key = field["key"]
            value = self.vars[key].get().strip()
            if field.get("required") and not value:
                messagebox.showerror("Missing value", f"{field['label']} is required.", parent=self)
                return
            payload[key] = value
        self.result = payload
        self.destroy()


class CrudWindow(tk.Toplevel):
    def __init__(self, parent, title: str, service: RcsDataService, config: dict):
        super().__init__(parent)
        self.title(title)
        self.geometry("1180x760")
        self.minsize(980, 640)
        self.configure(bg=OFF_WHITE)
        self.service = service
        self.config_data = config
        self.selected_id: int | None = None
        self.editing_id: int | None = None
        self.row_lookup: dict[int, dict] = {}
        self.search_vars: dict[str, tk.StringVar] = {}
        self.form_vars: dict[str, tk.StringVar] = {}
        self.form_widgets: dict[str, ttk.Entry | ttk.Combobox] = {}
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        controls = ttk.Frame(self, padding=14)
        controls.pack(fill="x")
        col = 0
        for field in self.config_data["search_fields"]:
            ttk.Label(controls, text=field["label"]).grid(row=0, column=col, sticky="w")
            var = tk.StringVar(value=field.get("default", ""))
            self.search_vars[field["key"]] = var
            if field.get("type") == "dropdown":
                widget = ttk.Combobox(controls, values=field["options"], textvariable=var, state="readonly", width=18)
                widget.grid(row=1, column=col, sticky="ew", padx=(0, 8))
            else:
                widget = ttk.Entry(controls, textvariable=var, width=18)
                widget.grid(row=1, column=col, sticky="ew", padx=(0, 6))
                if field.get("type") == "date":
                    ttk.Button(controls, text="Calendar", command=lambda v=var: self._pick_date(v)).grid(row=1, column=col + 1, padx=(0, 8))
                    col += 1
            col += 1
        ttk.Button(controls, text="Search", command=self.refresh).grid(row=1, column=col, padx=(6, 6))
        ttk.Button(controls, text="Add", command=self.add_record).grid(row=1, column=col + 1, padx=(0, 6))
        ttk.Button(controls, text="Edit", command=self.edit_record).grid(row=1, column=col + 2, padx=(0, 6))
        ttk.Button(controls, text="Delete", command=self.delete_record).grid(row=1, column=col + 3, padx=(0, 6))
        ttk.Button(controls, text="Back", command=self.destroy).grid(row=1, column=col + 4)

        form_card = ttk.LabelFrame(self, text="Row Editor", padding=14)
        form_card.pack(fill="x", padx=14, pady=(0, 8))
        row = 0
        for field in self.config_data["form_fields"]:
            key = field["key"]
            var = tk.StringVar(value=str(self.config_data.get("initial_defaults", {}).get(key, field.get("default", "")) or ""))
            self.form_vars[key] = var
            ttk.Label(form_card, text=field["label"]).grid(row=row, column=0, sticky="w", pady=(8 if row else 0, 4))
            if field.get("type") == "dropdown":
                widget = ttk.Combobox(form_card, values=field["options"], textvariable=var, state="readonly", width=26)
                widget.grid(row=row + 1, column=0, sticky="ew", padx=(0, 8))
            else:
                widget = ttk.Entry(form_card, textvariable=var, width=28, state="readonly" if field.get("readonly") else "normal")
                widget.grid(row=row + 1, column=0, sticky="ew", padx=(0, 8))
                if field.get("type") == "date":
                    ttk.Button(form_card, text="Calendar", command=lambda v=var: self._pick_date(v)).grid(row=row + 1, column=1, padx=(0, 8))
            self.form_widgets[key] = widget
            row += 2

        if "client" in self.form_vars and "value_ugx" in self.form_vars and "exchange_rate" in self.form_vars and "value_usd" in self.form_vars:
            self.form_vars["client"].trace_add("write", lambda *_: self._toggle_monthly_entry_mode())
            self.form_vars["value_ugx"].trace_add("write", lambda *_: self._recompute_monthly_values())
            self.form_vars["exchange_rate"].trace_add("write", lambda *_: self._recompute_monthly_values())
            self.form_vars["value_usd"].trace_add("write", lambda *_: self._recompute_monthly_values())
            self._toggle_monthly_entry_mode()

        actions = ttk.Frame(form_card)
        actions.grid(row=0, column=2, rowspan=max(2, row), sticky="ne", padx=(12, 0))
        self.editor_status = tk.StringVar(value="Ready to add a new row.")
        ttk.Label(actions, textvariable=self.editor_status, foreground=MUTED, wraplength=220, justify="left").pack(anchor="w", pady=(0, 10))
        ttk.Button(actions, text="Save Row", command=self.save_record).pack(fill="x", pady=(0, 6))
        ttk.Button(actions, text="Clear Editor", command=self.reset_editor).pack(fill="x")
        form_card.columnconfigure(0, weight=1)

        card = ttk.Frame(self, padding=14)
        card.pack(fill=BOTH, expand=True)
        self.tree = ttk.Treeview(card, columns=[c[0] for c in self.config_data["columns"]], show="headings", height=20)
        vsb = ttk.Scrollbar(card, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)
        self.column_labels = dict(self.config_data["columns"])
        for key, label in self.config_data["columns"]:
            self.tree.heading(key, text=label)
            anchor = "e" if key in {"value", "amount_paid", "value_ugx", "exchange_rate", "value_usd"} else "center"
            self.tree.column(key, width=130 if key != "comment" else 220, anchor=anchor)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-1>", lambda _event: self.edit_record())

    def _pick_date(self, var: tk.StringVar):
        selected = ask_date(self, var.get().strip() or None)
        if selected:
            var.set(selected)

    def _on_select(self, _event=None):
        selection = self.tree.selection()
        self.selected_id = int(selection[0]) if selection else None

    def _toggle_monthly_entry_mode(self):
        client = self.form_vars.get("client", tk.StringVar()).get().strip()
        ugx_widget = self.form_widgets.get("value_ugx")
        usd_widget = self.form_widgets.get("value_usd")
        if ugx_widget is None or usd_widget is None:
            return
        if client == "Totalenergies EP":
            ugx_widget.configure(state="readonly")
            usd_widget.configure(state="normal")
        else:
            ugx_widget.configure(state="normal")
            usd_widget.configure(state="readonly")
        self._recompute_monthly_values()

    def _recompute_monthly_values(self):
        if not {"client", "value_ugx", "exchange_rate", "value_usd"}.issubset(self.form_vars):
            return
        client = self.form_vars["client"].get().strip()
        rate = parse_number(self.form_vars["exchange_rate"].get(), 0.0) or 0.0
        try:
            if client == "Totalenergies EP":
                usd = parse_number(self.form_vars["value_usd"].get(), 0.0) or 0.0
                ugx = usd * rate if rate else 0.0
                self.form_vars["value_ugx"].set(f"{ugx:.2f}")
            else:
                ugx = parse_number(self.form_vars["value_ugx"].get(), 0.0) or 0.0
                usd = ugx / rate if rate else 0.0
                self.form_vars["value_usd"].set(f"{usd:.2f}")
        except ValueError:
            return

    def _refresh_dynamic_options(self):
        for field in self.config_data["form_fields"]:
            key = field["key"]
            widget = self.form_widgets.get(key)
            if widget is None or not isinstance(widget, ttk.Combobox):
                continue
            options = field.get("options")
            if field.get("options_method"):
                options = getattr(self.service, field["options_method"])()
            if options is not None:
                widget.configure(values=options)

    def _wire_dynamic_form_behaviors(self):
        if "voucher_type" in self.form_vars and "cost" in self.form_vars:
            self.form_vars["voucher_type"].trace_add("write", lambda *_: self._autofill_voucher_cost())
        if "iridium_number" in self.form_vars and "client" in self.form_vars:
            self.form_vars["iridium_number"].trace_add("write", lambda *_: self._autofill_sim_client())

    def _autofill_voucher_cost(self):
        if self._autofill_lock:
            return
        voucher_name = self.form_vars.get("voucher_type", tk.StringVar()).get().strip()
        if not voucher_name:
            return
        record = self.service.voucher_by_name(voucher_name)
        if not record:
            return
        self._autofill_lock = True
        try:
            self.form_vars["voucher_type"].set(record["name"])
            self.form_vars["cost"].set(f"{(parse_number(record['cost'], 0.0) or 0.0):.2f}")
        finally:
            self._autofill_lock = False

    def _autofill_sim_client(self):
        if self._autofill_lock:
            return
        typed_value = self.form_vars.get("iridium_number", tk.StringVar()).get().strip()
        if not typed_value:
            return
        record = self.service.match_sim_msisdn(typed_value)
        if not record:
            return
        self._autofill_lock = True
        try:
            self.form_vars["iridium_number"].set(record["msisdn"])
            self.form_vars["client"].set(record["client"])
        finally:
            self._autofill_lock = False

    def import_iridium_sims_excel(self):
        file_path = filedialog.askopenfilename(
            parent=self,
            title="Select Iridium SIM import file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return
        try:
            result = self.service.import_iridium_sims_excel(file_path)
        except Exception as exc:
            messagebox.showerror(self.config_data["title"], str(exc), parent=self)
            return
        self.refresh()
        messagebox.showinfo(
            self.config_data["title"],
            f"Iridium SIM import completed.\nInserted: {result['inserted']}\nUpdated: {result['updated']}",
            parent=self,
        )

    def import_marlink_excel(self):
        file_path = filedialog.askopenfilename(
            parent=self,
            title="Select Marlink import file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return
        try:
            result = self.service.import_marlink_excel(file_path)
        except Exception as exc:
            messagebox.showerror(self.config_data["title"], str(exc), parent=self)
            return
        self.refresh()
        messagebox.showinfo(
            self.config_data["title"],
            f"Marlink import completed.\nInserted: {result['inserted']}\nUpdated: {result['updated']}",
            parent=self,
        )

    def reset_editor(self):
        self.editing_id = None
        self._refresh_dynamic_options()
        for field in self.config_data["form_fields"]:
            key = field["key"]
            default = self.config_data.get("initial_defaults", {}).get(key, field.get("default", ""))
            self.form_vars[key].set(str(default or ""))
        if "client" in self.form_vars:
            self._toggle_monthly_entry_mode()
        self.editor_status.set("Ready to add a new row.")

    def refresh(self):
        self.tree.delete(*self.tree.get_children())
        filters = {key: var.get() for key, var in self.search_vars.items()}
        filters.update(self.config_data.get("fixed_search", {}))
        rows = getattr(self.service, self.config_data["list_method"])(**filters)
        self.row_lookup = {}
        totals: dict[str, float] = {key: 0.0 for key in self.config_data.get("total_columns", [])}
        for row in rows:
            row_dict = dict(row)
            self.row_lookup[int(row_dict["id"])] = row_dict
            values = []
            for key, _label in self.config_data["columns"]:
                value = row[key]
                if key in totals:
                    totals[key] += parse_number(row[key], 0.0) or 0.0
                if key in self.config_data.get("usd_columns", []):
                    value = format_usd(value)
                elif key in self.config_data.get("ugx_columns", []):
                    value = format_ugx(value)
                elif key in {"value", "amount_paid", "value_ugx", "exchange_rate", "value_usd"}:
                    value = format_amount(value)
                values.append(value)
            self.tree.insert("", END, values=tuple(values))
        for key, label in self.column_labels.items():
            if key in self.config_data.get("usd_columns", []):
                self.tree.heading(key, text=f"{label}: {format_usd(totals.get(key, 0.0))}" if rows else label)
            elif key in self.config_data.get("ugx_columns", []):
                self.tree.heading(key, text=f"{label}: {format_ugx(totals.get(key, 0.0))}" if rows else label)
            elif key in totals:
                self.tree.heading(key, text=f"{label}: {format_amount(totals.get(key, 0.0))}" if rows else label)
            else:
                self.tree.heading(key, text=label)
        self.selected_id = None

    def add_record(self):
        self.reset_editor()
        self.editor_status.set("Adding a new row. Fill the editor and click Save Row.")

    def edit_record(self):
        if self.selected_id is None:
            messagebox.showwarning(self.config_data["title"], "Please select a row first.", parent=self)
            return
        row_map = self.row_lookup.get(self.selected_id)
        if not row_map:
            messagebox.showerror(self.config_data["title"], "Could not load the selected row.", parent=self)
            return
        for field in self.config_data["form_fields"]:
            key = field["key"]
            raw = row_map.get(key, "")
            if raw in (None, ""):
                raw = ""
            if key in self.config_data.get("usd_columns", []):
                raw = str(raw).replace("$", "").replace(",", "")
            elif key in self.config_data.get("ugx_columns", []):
                raw = str(raw).replace("/=", "").replace(",", "")
            elif key in {"exchange_rate", "units", "cost", "marlink_invoice_value"}:
                raw = str(raw).replace(",", "")
            self.form_vars[key].set(str(raw))
        self.editing_id = self.selected_id
        if "client" in self.form_vars:
            self._toggle_monthly_entry_mode()
        self.editor_status.set(f"Editing row ID {self.editing_id}. Update the values and click Save Row.")

    def save_record(self):
        payload = {}
        for field in self.config_data["form_fields"]:
            key = field["key"]
            value = self.form_vars[key].get().strip()
            if field.get("required"):
                if key == "value_ugx" and "client" in self.form_vars and self.form_vars["client"].get().strip() == "Totalenergies EP":
                    pass
                elif not value:
                    messagebox.showerror(self.config_data["title"], f"{field['label']} is required.", parent=self)
                    return
            if key in {"value", "amount_paid", "exchange_rate", "value_usd", "value_ugx", "units", "cost", "marlink_invoice_value"} and value:
                parsed = parse_number(value, None)
                value = "" if parsed is None else str(parsed)
            payload[key] = value
        try:
            if self.editing_id is None:
                getattr(self.service, self.config_data["create_method"])(payload)
            else:
                getattr(self.service, self.config_data["update_method"])(self.editing_id, payload)
        except Exception as exc:
            messagebox.showerror(self.config_data["title"], str(exc), parent=self)
            return
        self.refresh()
        self.reset_editor()

    def delete_record(self):
        if self.selected_id is None:
            messagebox.showwarning(self.config_data["title"], "Please select a row first.", parent=self)
            return
        if not messagebox.askyesno(self.config_data["title"], "Delete the selected row?", parent=self):
            return
        getattr(self.service, self.config_data["delete_method"])(self.selected_id)
        self.refresh()


class LedgerWindow(tk.Toplevel):
    def __init__(self, parent, title: str, data: dict):
        super().__init__(parent)
        self.title(title)
        self.geometry("1120x720")
        self.minsize(960, 620)
        self.configure(bg=OFF_WHITE)
        totals = ttk.Frame(self, padding=14)
        totals.pack(fill="x")
        ttk.Button(totals, text="Back", command=self.destroy).pack(side="left", padx=(0, 16))
        ttk.Label(totals, text=f"Total invoices: {format_usd(data['total_invoices'])}", font=("Segoe UI Semibold", 11)).pack(side="left", padx=(0, 16))
        ttk.Label(totals, text=f"Total payments: {format_usd(data['total_payments'])}", font=("Segoe UI Semibold", 11)).pack(side="left", padx=(0, 16))
        ttk.Label(totals, text=f"Balance: {format_usd(data['balance'])}", font=("Segoe UI Semibold", 11), foreground=NEGATIVE if data['balance'] < 0 else POSITIVE).pack(side="left")

        card = ttk.Frame(self, padding=14)
        card.pack(fill=BOTH, expand=True)
        tree = ttk.Treeview(card, columns=("invoice_date", "supplier", "reference", "invoice_value", "payment_date", "payment_value", "balance"), show="headings")
        vsb = ttk.Scrollbar(card, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)
        headings = {
            "invoice_date": "Invoice Date",
            "supplier": "Supplier",
            "reference": "Reference",
            "invoice_value": f"Value: {format_usd(data['total_invoices'])}",
            "payment_date": "Payment Date",
            "payment_value": f"Value: {format_usd(data['total_payments'])}",
            "balance": "Balance",
        }
        for key, label in headings.items():
            tree.heading(key, text=label)
            anchor = "e" if key in {"invoice_value", "payment_value", "balance"} else "center"
            tree.column(key, width=140, anchor=anchor)
        tree.tag_configure("negative", foreground=NEGATIVE)
        tree.tag_configure("positive", foreground=POSITIVE)
        for row in data["rows"]:
            tag = "negative" if row["balance"] < 0 else "positive"
            tree.insert(
                "",
                END,
                values=(
                    row["invoice_date"],
                    row["supplier"],
                    row["reference"],
                    format_usd(row["invoice_value"]) if row["invoice_value"] else "",
                    row["payment_date"],
                    format_usd(row["payment_value"]) if row["payment_value"] else "",
                    format_usd(row["balance"]),
                ),
                tags=(tag,),
            )


class CrudScreen(ttk.Frame):
    def __init__(self, parent, title: str, service: RcsDataService, config: dict, on_back):
        super().__init__(parent, padding=14)
        self.title_text = title
        self.service = service
        self.config_data = config
        self.on_back = on_back
        self.selected_id: int | None = None
        self.editing_id: int | None = None
        self.row_lookup: dict[int, dict] = {}
        self.search_vars: dict[str, tk.StringVar] = {}
        self.form_vars: dict[str, tk.StringVar] = {}
        self.form_widgets: dict[str, ttk.Entry | ttk.Combobox] = {}
        self.tree = None
        self.group_canvas = None
        self.group_frame = None
        self.group_row_widgets: dict[int, list[tk.Widget]] = {}
        self._autofill_lock = False
        self.commission_var = tk.StringVar(value=str(self.config_data.get("default_commission_pct", 10)))
        self.pack(fill=BOTH, expand=True)
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        header = ttk.Frame(self)
        header.pack(fill="x", pady=(0, 8))
        ttk.Button(header, text="Back", command=self.on_back).pack(side="left", padx=(0, 12))
        ttk.Label(header, text=self.title_text, font=("Segoe UI Semibold", 18), foreground=NAVY).pack(side="left")

        self.summary_frame = ttk.Frame(self)
        if self.config_data.get("summary_mode"):
            self.summary_frame.pack(fill="x", pady=(0, 8))
        self.summary_vars = {
            "invoiced": tk.StringVar(value=format_usd(0)),
            "paid": tk.StringVar(value=format_usd(0)),
            "balance": tk.StringVar(value=format_usd(0)),
            "vat_incl": tk.StringVar(value=format_usd(0)),
            "vat_ex": tk.StringVar(value=format_usd(0)),
            "commission": tk.StringVar(value=format_usd(0)),
            "kla_cost": tk.StringVar(value=format_usd(0)),
        }
        self.balance_label = None
        if self.config_data.get("summary_mode"):
            self.commission_var.trace_add("write", lambda *_: self._update_summary_strip(getattr(self, "_current_rows", [])))
            self._build_summary_strip()

        controls = ttk.Frame(self)
        controls.pack(fill="x", pady=(0, 6))
        col = 0
        for field in self.config_data["search_fields"]:
            ttk.Label(controls, text=field["label"]).grid(row=0, column=col, sticky="w")
            var = tk.StringVar(value=field.get("default", ""))
            self.search_vars[field["key"]] = var
            if field.get("type") == "dropdown":
                widget = ttk.Combobox(controls, values=field["options"], textvariable=var, state="readonly", width=18)
                widget.grid(row=1, column=col, sticky="ew", padx=(0, 8))
            else:
                widget = ttk.Entry(controls, textvariable=var, width=18)
                widget.grid(row=1, column=col, sticky="ew", padx=(0, 6))
                if field.get("type") == "date":
                    ttk.Button(controls, text="Calendar", command=lambda v=var: self._pick_date(v)).grid(row=1, column=col + 1, padx=(0, 8))
                    col += 1
            col += 1
        button_row = 2
        ttk.Button(controls, text="Search", command=self.refresh).grid(row=button_row, column=0, padx=(0, 6), pady=(8, 0), sticky="w")
        ttk.Button(controls, text="Add", command=self.add_record).grid(row=button_row, column=1, padx=(0, 6), pady=(8, 0), sticky="w")
        ttk.Button(controls, text="Edit", command=self.edit_record).grid(row=button_row, column=2, padx=(0, 6), pady=(8, 0), sticky="w")
        ttk.Button(controls, text="Delete", command=self.delete_record).grid(row=button_row, column=3, padx=(0, 6), pady=(8, 0), sticky="w")
        action_col = 4
        for action in self.config_data.get("extra_actions", []):
            ttk.Button(controls, text=action["label"], command=getattr(self, action["command"])).grid(
                row=button_row, column=action_col, padx=(0, 6), pady=(8, 0), sticky="w"
            )
            action_col += 1
        ttk.Button(controls, text="Back", command=self.on_back).grid(row=button_row, column=action_col, pady=(8, 0), sticky="w")
        for idx in range(max(col, action_col) + 1):
            controls.columnconfigure(idx, weight=1 if idx < col else 0)

        form_card = ttk.LabelFrame(self, text="Row Editor", padding=8)
        form_card.pack(fill="x", pady=(6, 6))
        field_rows = max(1, (len(self.config_data["form_fields"]) + 1) // 2)
        for idx, field in enumerate(self.config_data["form_fields"]):
            key = field["key"]
            var = tk.StringVar(value=str(self.config_data.get("initial_defaults", {}).get(key, field.get("default", "")) or ""))
            self.form_vars[key] = var
            row = idx // 2
            col_block = (idx % 2) * 3
            ttk.Label(form_card, text=field["label"]).grid(row=row, column=col_block, sticky="w", padx=(0, 6), pady=2)
            if field.get("type") == "dropdown":
                widget = ttk.Combobox(
                    form_card,
                    values=field.get("options", []),
                    textvariable=var,
                    state=field.get("combobox_state", "readonly"),
                    width=20,
                )
                widget.grid(row=row, column=col_block + 1, sticky="ew", padx=(0, 8), pady=2)
            else:
                widget = ttk.Entry(form_card, textvariable=var, width=22, state="readonly" if field.get("readonly") else "normal")
                widget.grid(row=row, column=col_block + 1, sticky="ew", padx=(0, 8), pady=2)
                if field.get("type") == "date":
                    ttk.Button(form_card, text="Cal", width=5, command=lambda v=var: self._pick_date(v)).grid(row=row, column=col_block + 2, padx=(0, 10), pady=2)
            self.form_widgets[key] = widget

        if "client" in self.form_vars and "value_ugx" in self.form_vars and "exchange_rate" in self.form_vars and "value_usd" in self.form_vars:
            self.form_vars["client"].trace_add("write", lambda *_: self._toggle_monthly_entry_mode())
            self.form_vars["value_ugx"].trace_add("write", lambda *_: self._recompute_monthly_values())
            self.form_vars["exchange_rate"].trace_add("write", lambda *_: self._recompute_monthly_values())
            self.form_vars["value_usd"].trace_add("write", lambda *_: self._recompute_monthly_values())
            self._toggle_monthly_entry_mode()
        self._refresh_dynamic_options()
        self._wire_dynamic_form_behaviors()

        actions = ttk.Frame(form_card)
        actions.grid(row=0, column=6, rowspan=max(1, field_rows), sticky="ne", padx=(10, 0))
        self.editor_status = tk.StringVar(value="Ready to add a new row.")
        ttk.Label(actions, textvariable=self.editor_status, foreground=MUTED, wraplength=180, justify="left").pack(anchor="w", pady=(0, 6))
        ttk.Button(actions, text="Save Row", command=self.save_record).pack(fill="x", pady=(0, 6))
        ttk.Button(actions, text="Clear Editor", command=self.reset_editor).pack(fill="x")
        form_card.columnconfigure(1, weight=1)
        form_card.columnconfigure(4, weight=1)

        card = ttk.Frame(self, padding=8)
        card.pack(fill=BOTH, expand=True)
        style = ttk.Style()
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)
        self.column_labels = dict(self.config_data["columns"])
        if self.config_data.get("group_by_month"):
            self.group_canvas = tk.Canvas(card, bg="white", highlightthickness=0)
            vsb = ttk.Scrollbar(card, orient=VERTICAL, command=self.group_canvas.yview)
            self.group_canvas.configure(yscrollcommand=vsb.set)
            self.group_canvas.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            self.group_frame = tk.Frame(self.group_canvas, bg="white")
            canvas_window = self.group_canvas.create_window((0, 0), window=self.group_frame, anchor="nw")

            def _sync_scrollregion(_event=None):
                self.group_canvas.configure(scrollregion=self.group_canvas.bbox("all"))

            def _sync_canvas_width(event):
                self.group_canvas.itemconfigure(canvas_window, width=event.width)

            self.group_frame.bind("<Configure>", _sync_scrollregion)
            self.group_canvas.bind("<Configure>", _sync_canvas_width)
        else:
            tree_style = f"{self.title_text}.Treeview"
            heading_style = f"{self.title_text}.Treeview.Heading"
            style.configure(tree_style, rowheight=22)
            style.configure(heading_style, font=("Segoe UI Semibold", 9))
            self.tree = ttk.Treeview(card, columns=[c[0] for c in self.config_data["columns"]], show="headings", height=20, style=tree_style)
            vsb = ttk.Scrollbar(card, orient=VERTICAL, command=self.tree.yview)
            self.tree.configure(yscrollcommand=vsb.set)
            self.tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            for key, label in self.config_data["columns"]:
                self.tree.heading(key, text=label)
                anchor = "e" if key in {"value", "amount_paid", "value_ugx", "exchange_rate", "value_usd", "cost", "marlink_invoice_value"} else "center"
                self.tree.column(key, width=130 if key != "comment" else 220, anchor=anchor)
            self.tree.bind("<<TreeviewSelect>>", self._on_select)
            self.tree.bind("<Double-1>", lambda _event: self.edit_record())

    def _build_summary_strip(self):
        if self.config_data.get("summary_mode") == "monthly_invoice":
            items = [
                ("Total invoiced VAT Incl", self.summary_vars["vat_incl"], False),
                ("Total invoiced VAT Ex", self.summary_vars["vat_ex"], False),
                ("Kla Commission", self.commission_var, True),
                ("Commission", self.summary_vars["commission"], False),
                ("Kla Cost", self.summary_vars["kla_cost"], False),
            ]
        else:
            items = [
                ("Total invoiced", self.summary_vars["invoiced"], False),
                ("Total paid", self.summary_vars["paid"], False),
                ("Balance", self.summary_vars["balance"], False),
            ]
        for idx, (label, var, editable) in enumerate(items):
            card = tk.Frame(self.summary_frame, bg="white", highlightbackground="#d7e2ea", highlightthickness=1, padx=10, pady=7)
            card.grid(row=0, column=idx, sticky="nsew", padx=(0, 8) if idx < len(items) - 1 else 0)
            line = tk.Frame(card, bg="white")
            line.pack(fill="x")
            tk.Label(line, text=f"{label}:", bg="white", fg=MUTED, font=("Segoe UI Semibold", 9)).pack(side="left")
            if editable:
                ttk.Entry(line, textvariable=var, width=6).pack(side="left", padx=(8, 4))
                tk.Label(line, text="%", bg="white", fg=MUTED, font=("Segoe UI Semibold", 9)).pack(side="left")
                value_label = None
            else:
                value_label = tk.Label(line, textvariable=var, bg="white", fg=TEXT_DARK, font=("Segoe UI Semibold", 11))
                value_label.pack(side="left", padx=(8, 0))
            if label == "Balance":
                self.balance_label = value_label
        for idx in range(len(items)):
            self.summary_frame.columnconfigure(idx, weight=1)

    def _update_summary_strip(self, rows):
        if not self.config_data.get("summary_mode"):
            return
        total_invoiced = 0.0
        total_paid = 0.0
        mode = self.config_data.get("summary_mode")
        for row in rows:
            if mode == "invoice_amount_paid":
                total_invoiced += parse_number(row["value"], 0.0) or 0.0
                total_paid += parse_number(row["amount_paid"], 0.0) or 0.0
            elif mode == "monthly_invoice":
                total_invoiced += parse_number(row["value_usd"], 0.0) or 0.0
        if mode == "monthly_invoice":
            vat_incl = total_invoiced
            vat_ex = vat_incl / 1.18 if vat_incl else 0.0
            commission_pct = (parse_number(self.commission_var.get(), self.config_data.get("default_commission_pct", 10)) or 0.0) / 100.0
            commission_value = vat_ex * commission_pct
            kla_cost = vat_ex - commission_value
            self.summary_vars["vat_incl"].set(format_usd(vat_incl))
            self.summary_vars["vat_ex"].set(format_usd(vat_ex))
            self.summary_vars["commission"].set(format_usd(commission_value))
            self.summary_vars["kla_cost"].set(format_usd(kla_cost))
        else:
            balance = total_paid - total_invoiced
            self.summary_vars["invoiced"].set(format_usd(total_invoiced))
            self.summary_vars["paid"].set(format_usd(total_paid))
            self.summary_vars["balance"].set(format_usd(balance))
            if self.balance_label is not None:
                color = POSITIVE if balance > 0 else NEGATIVE if balance < 0 else TEXT_DARK
                self.balance_label.configure(fg=color)

    def _pick_date(self, var: tk.StringVar):
        selected = ask_date(self, var.get().strip() or None)
        if selected:
            var.set(selected)

    def _on_select(self, _event=None):
        selection = self.tree.selection()
        self.selected_id = int(selection[0]) if selection else None

    def _select_group_row(self, row_id: int):
        self.selected_id = row_id
        selected_bg = "#dbefff"
        default_bg = "#ffffff"
        for item_id, widgets in self.group_row_widgets.items():
            for widget in widgets:
                try:
                    widget.configure(bg=selected_bg if item_id == row_id else default_bg)
                except tk.TclError:
                    pass

    def _render_grouped_rows(self, rows):
        if self.group_frame is None:
            return
        for child in self.group_frame.winfo_children():
            child.destroy()
        self.group_row_widgets = {}

        width_overrides = self.config_data.get("group_column_widths", {})

        def cell_width_for(key: str) -> int:
            if key in width_overrides:
                return width_overrides[key]
            if key in {"client", "voucher_type"}:
                return 20
            if key in {"invoice_number", "marlink_invoice"}:
                return 18
            return 16

        header = tk.Frame(self.group_frame, bg="#eef4f7")
        header.pack(fill="x", pady=(0, 8))
        for idx, (key, label) in enumerate(self.config_data["columns"]):
            anchor = "e" if key in {"cost", "marlink_invoice_value"} else "center"
            width = cell_width_for(key)
            tk.Label(
                header,
                text=label,
                bg="#eef4f7",
                fg=TEXT_DARK,
                font=("Segoe UI Semibold", 9),
                anchor=anchor,
                padx=6,
                pady=5,
                relief="solid",
                bd=1,
                width=width,
            ).grid(row=0, column=idx, sticky="nsew")
            header.columnconfigure(idx, weight=1)

        groups: dict[str, list[dict]] = {}
        date_key = self.config_data.get("group_date_key", "entry_date")
        for row in rows:
            row_dict = dict(row)
            month_key = (row_dict.get(date_key) or "")[:7] or "No month"
            groups.setdefault(month_key, []).append(row_dict)

        for month_key, month_rows in groups.items():
            outer = tk.Frame(self.group_frame, bg="#d8e2ea", padx=1, pady=1)
            outer.pack(fill="x", pady=(0, 10))
            inner = tk.Frame(outer, bg="white")
            inner.pack(fill="x")

            month_bar = tk.Frame(inner, bg="#edf3f8", padx=10, pady=6)
            month_bar.pack(fill="x")
            tk.Label(month_bar, text=month_key, bg="#edf3f8", fg=NAVY, font=("Segoe UI Semibold", 11)).pack(side="left")

            group_total_keys = self.config_data.get("group_total_columns") or self.config_data.get("total_columns", [])
            group_totals = {key: 0.0 for key in group_total_keys}
            for row_dict in month_rows:
                row_id = int(row_dict["id"])
                for key in group_total_keys:
                    group_totals[key] += parse_number(row_dict.get(key), 0.0) or 0.0
                row_frame = tk.Frame(inner, bg="white")
                row_frame.pack(fill="x")
                widgets = [row_frame]
                for idx, (key, _label) in enumerate(self.config_data["columns"]):
                    value = row_dict.get(key, "")
                    if key in self.config_data.get("usd_columns", []):
                        value = format_usd(value)
                    elif key in self.config_data.get("ugx_columns", []):
                        value = format_ugx(value)
                    anchor = "e" if key in {"cost", "marlink_invoice_value"} else "center"
                    width = cell_width_for(key)
                    cell = tk.Label(
                        row_frame,
                        text=value,
                        bg="white",
                        fg=TEXT_DARK,
                        font=("Segoe UI", 9),
                        anchor=anchor,
                        padx=6,
                        pady=4,
                        relief="solid",
                        bd=1,
                        width=width,
                    )
                    cell.grid(row=0, column=idx, sticky="nsew")
                    row_frame.columnconfigure(idx, weight=1)
                    widgets.append(cell)
                for widget in widgets:
                    widget.bind("<Button-1>", lambda _event, rid=row_id: self._select_group_row(rid))
                    widget.bind("<Double-1>", lambda _event, rid=row_id: (self._select_group_row(rid), self.edit_record()))
                self.group_row_widgets[row_id] = widgets

            total_bar = tk.Frame(inner, bg="white", padx=10, pady=8)
            total_bar.pack(fill="x")
            total_box = tk.Frame(total_bar, bg="#f3f7fb", highlightbackground="#b8c8d5", highlightthickness=1, padx=10, pady=6)
            total_box.pack(side="right")
            total_text = self._build_group_total_text(month_key, group_totals)
            tk.Label(total_box, text=total_text, bg="#f3f7fb", fg=TEXT_DARK, font=("Segoe UI Semibold", 9)).pack()

    def _build_group_total_text(self, month_key: str, group_totals: dict[str, float]) -> str:
        if self.config_data.get("group_total_mode") == "monthly_invoice":
            vat_incl = group_totals.get("value_usd", 0.0)
            vat_ex = vat_incl / 1.18 if vat_incl else 0.0
            commission_pct = (parse_number(self.commission_var.get(), self.config_data.get("default_commission_pct", 10)) or 0.0) / 100.0
            commission_value = vat_ex * commission_pct
            net_total = vat_ex - commission_value
            return (
                f"Total {month_key} = "
                f"VAT Incl {format_usd(vat_incl)} | "
                f"VAT Ex {format_usd(vat_ex)} | "
                f"Commission {format_usd(commission_value)} | "
                f"VAT Ex - Commission {format_usd(net_total)}"
            )

        total_parts = []
        for key in self.config_data.get("group_total_columns") or self.config_data.get("total_columns", []):
            label = self.column_labels.get(key, key)
            total_value = group_totals.get(key, 0.0)
            if key in self.config_data.get("usd_columns", []):
                formatted = format_usd(total_value)
            elif key in self.config_data.get("ugx_columns", []):
                formatted = format_ugx(total_value)
            else:
                formatted = format_amount(total_value)
            total_parts.append(f"{label} {formatted}")
        return f"Total {month_key} = {' | '.join(total_parts)}" if total_parts else f"Total {month_key}"

    def _toggle_monthly_entry_mode(self):
        client = self.form_vars.get("client", tk.StringVar()).get().strip()
        ugx_widget = self.form_widgets.get("value_ugx")
        usd_widget = self.form_widgets.get("value_usd")
        if ugx_widget is None or usd_widget is None:
            return
        if client == "Totalenergies EP":
            ugx_widget.configure(state="readonly")
            usd_widget.configure(state="normal")
        else:
            ugx_widget.configure(state="normal")
            usd_widget.configure(state="readonly")
        self._recompute_monthly_values()

    def _recompute_monthly_values(self):
        if not {"client", "value_ugx", "exchange_rate", "value_usd"}.issubset(self.form_vars):
            return
        client = self.form_vars["client"].get().strip()
        rate = parse_number(self.form_vars["exchange_rate"].get(), 0.0) or 0.0
        try:
            if client == "Totalenergies EP":
                usd = parse_number(self.form_vars["value_usd"].get(), 0.0) or 0.0
                ugx = usd * rate if rate else 0.0
                self.form_vars["value_ugx"].set(f"{ugx:.2f}")
            else:
                ugx = parse_number(self.form_vars["value_ugx"].get(), 0.0) or 0.0
                usd = ugx / rate if rate else 0.0
                self.form_vars["value_usd"].set(f"{usd:.2f}")
        except ValueError:
            return

    def _refresh_dynamic_options(self):
        for field in self.config_data["form_fields"]:
            key = field["key"]
            widget = self.form_widgets.get(key)
            if widget is None or not isinstance(widget, ttk.Combobox):
                continue
            options = field.get("options")
            if field.get("options_method"):
                options = getattr(self.service, field["options_method"])()
            if options is not None:
                widget.configure(values=options)

    def _wire_dynamic_form_behaviors(self):
        if "voucher_type" in self.form_vars and "cost" in self.form_vars:
            self.form_vars["voucher_type"].trace_add("write", lambda *_: self._autofill_voucher_cost())
        if "iridium_number" in self.form_vars and "client" in self.form_vars:
            self.form_vars["iridium_number"].trace_add("write", lambda *_: self._autofill_sim_client())

    def _autofill_voucher_cost(self):
        if self._autofill_lock:
            return
        voucher_name = self.form_vars.get("voucher_type", tk.StringVar()).get().strip()
        if not voucher_name:
            return
        record = self.service.voucher_by_name(voucher_name)
        if not record:
            return
        self._autofill_lock = True
        try:
            self.form_vars["voucher_type"].set(record["name"])
            self.form_vars["cost"].set(f"{(parse_number(record['cost'], 0.0) or 0.0):.2f}")
        finally:
            self._autofill_lock = False

    def _autofill_sim_client(self):
        if self._autofill_lock:
            return
        typed_value = self.form_vars.get("iridium_number", tk.StringVar()).get().strip()
        if not typed_value:
            return
        record = self.service.match_sim_msisdn(typed_value)
        if not record:
            return
        self._autofill_lock = True
        try:
            self.form_vars["iridium_number"].set(record["msisdn"])
            self.form_vars["client"].set(record["client"])
        finally:
            self._autofill_lock = False

    def import_iridium_sims_excel(self):
        file_path = filedialog.askopenfilename(
            parent=self,
            title="Select Iridium SIM import file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return
        try:
            result = self.service.import_iridium_sims_excel(file_path)
        except Exception as exc:
            messagebox.showerror(self.config_data["title"], str(exc), parent=self)
            return
        self.refresh()
        messagebox.showinfo(
            self.config_data["title"],
            f"Iridium SIM import completed.\nInserted: {result['inserted']}\nUpdated: {result['updated']}",
            parent=self,
        )

    def import_marlink_excel(self):
        file_path = filedialog.askopenfilename(
            parent=self,
            title="Select Marlink import file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not file_path:
            return
        try:
            result = self.service.import_marlink_excel(file_path)
        except Exception as exc:
            messagebox.showerror(self.config_data["title"], str(exc), parent=self)
            return
        self.refresh()
        messagebox.showinfo(
            self.config_data["title"],
            f"Marlink import completed.\nInserted: {result['inserted']}\nUpdated: {result['updated']}",
            parent=self,
        )

    def reset_editor(self):
        self.editing_id = None
        self._refresh_dynamic_options()
        for field in self.config_data["form_fields"]:
            key = field["key"]
            default = self.config_data.get("initial_defaults", {}).get(key, field.get("default", ""))
            self.form_vars[key].set(str(default or ""))
        if "client" in self.form_vars:
            self._toggle_monthly_entry_mode()
        self.editor_status.set("Ready to add a new row.")

    def refresh(self):
        if self.tree is not None:
            self.tree.delete(*self.tree.get_children())
        self._refresh_dynamic_options()
        filters = {key: var.get() for key, var in self.search_vars.items()}
        filters.update(self.config_data.get("fixed_search", {}))
        rows = getattr(self.service, self.config_data["list_method"])(**filters)
        self._current_rows = rows
        self._update_summary_strip(rows)
        self.row_lookup = {}
        totals: dict[str, float] = {key: 0.0 for key in self.config_data.get("total_columns", [])}
        for row in rows:
            row_dict = dict(row)
            self.row_lookup[int(row_dict["id"])] = row_dict
            for key, _label in self.config_data["columns"]:
                if key in totals:
                    totals[key] += parse_number(row[key], 0.0) or 0.0
            if self.tree is not None:
                values = []
                for key, _label in self.config_data["columns"]:
                    value = row[key]
                    if key in self.config_data.get("usd_columns", []):
                        value = format_usd(value)
                    elif key in self.config_data.get("ugx_columns", []):
                        value = format_ugx(value)
                    elif key in {"value", "amount_paid", "value_ugx", "exchange_rate", "value_usd", "units", "cost", "marlink_invoice_value"}:
                        value = format_amount(value)
                    values.append(value)
                self.tree.insert("", END, iid=str(row_dict["id"]), values=tuple(values))
        if self.tree is not None:
            for key, label in self.column_labels.items():
                if key in self.config_data.get("usd_columns", []):
                    self.tree.heading(key, text=f"{label}: {format_usd(totals.get(key, 0.0))}" if rows else label)
                elif key in self.config_data.get("ugx_columns", []):
                    self.tree.heading(key, text=f"{label}: {format_ugx(totals.get(key, 0.0))}" if rows else label)
                elif key in totals:
                    self.tree.heading(key, text=f"{label}: {format_amount(totals.get(key, 0.0))}" if rows else label)
                else:
                    self.tree.heading(key, text=label)
        else:
            self._render_grouped_rows(rows)
        self.selected_id = None

    def add_record(self):
        self.reset_editor()
        self.editor_status.set("Adding a new row. Fill the editor and click Save Row.")

    def edit_record(self):
        if self.selected_id is None:
            selection = self.tree.selection() if self.tree is not None else ()
            self.selected_id = int(selection[0]) if selection else None
        if self.selected_id is None:
            messagebox.showwarning(self.config_data["title"], "Please select a row first.", parent=self)
            return
        row_map = self.row_lookup.get(self.selected_id)
        if not row_map:
            messagebox.showerror(self.config_data["title"], "Could not load the selected row.", parent=self)
            return
        for field in self.config_data["form_fields"]:
            key = field["key"]
            raw = row_map.get(key, "")
            if raw in (None, ""):
                raw = ""
            if key in self.config_data.get("usd_columns", []):
                raw = str(raw).replace("$", "").replace(",", "")
            elif key in self.config_data.get("ugx_columns", []):
                raw = str(raw).replace("/=", "").replace(",", "")
            elif key in {"exchange_rate", "units", "cost", "marlink_invoice_value"}:
                raw = str(raw).replace(",", "")
            self.form_vars[key].set(str(raw))
        self.editing_id = self.selected_id
        if "client" in self.form_vars:
            self._toggle_monthly_entry_mode()
        self.editor_status.set(f"Editing row ID {self.editing_id}. Update the values and click Save Row.")

    def save_record(self):
        payload = {}
        for field in self.config_data["form_fields"]:
            key = field["key"]
            value = self.form_vars[key].get().strip()
            if field.get("required"):
                if key == "value_ugx" and "client" in self.form_vars and self.form_vars["client"].get().strip() == "Totalenergies EP":
                    pass
                elif not value:
                    messagebox.showerror(self.config_data["title"], f"{field['label']} is required.", parent=self)
                    return
            if key in {"value", "amount_paid", "exchange_rate", "value_usd", "value_ugx", "units", "cost", "marlink_invoice_value"} and value:
                parsed = parse_number(value, None)
                value = "" if parsed is None else str(parsed)
            payload[key] = value
        try:
            if self.editing_id is None:
                getattr(self.service, self.config_data["create_method"])(payload)
            else:
                getattr(self.service, self.config_data["update_method"])(self.editing_id, payload)
        except Exception as exc:
            messagebox.showerror(self.config_data["title"], str(exc), parent=self)
            return
        self.refresh()
        self.reset_editor()

    def delete_record(self):
        if self.selected_id is None:
            selection = self.tree.selection() if self.tree is not None else ()
            self.selected_id = int(selection[0]) if selection else None
        if self.selected_id is None:
            messagebox.showwarning(self.config_data["title"], "Please select a row first.", parent=self)
            return
        if not messagebox.askyesno(self.config_data["title"], "Delete the selected row?", parent=self):
            return
        getattr(self.service, self.config_data["delete_method"])(self.selected_id)
        self.refresh()


class LedgerScreen(ttk.Frame):
    def __init__(self, parent, title: str, data: dict, on_back):
        super().__init__(parent, padding=10)
        self.pack(fill=BOTH, expand=True)
        header = ttk.Frame(self)
        header.pack(fill="x", pady=(0, 8))
        ttk.Button(header, text="Back", command=on_back).pack(side="left", padx=(0, 16))
        ttk.Label(header, text=title, font=("Segoe UI Semibold", 18), foreground=NAVY).pack(side="left")

        totals = ttk.Frame(self)
        totals.pack(fill="x", pady=(0, 8))
        ttk.Label(totals, text=f"Total invoices: {format_usd(data['total_invoices'])}", font=("Segoe UI Semibold", 10)).pack(side="left", padx=(0, 14))
        ttk.Label(totals, text=f"Total payments: {format_usd(data['total_payments'])}", font=("Segoe UI Semibold", 10)).pack(side="left", padx=(0, 14))
        ttk.Label(totals, text=f"Balance: {format_usd(data['balance'])}", font=("Segoe UI Semibold", 10), foreground=POSITIVE if data['balance'] < 0 else NEGATIVE if data['balance'] > 0 else TEXT_DARK).pack(side="left")

        card = ttk.Frame(self, padding=6)
        card.pack(fill=BOTH, expand=True)
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)
        canvas = tk.Canvas(card, bg="white", highlightthickness=0)
        vsb = ttk.Scrollbar(card, orient=VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        grid_frame = tk.Frame(canvas, bg="white")
        canvas_window = canvas.create_window((0, 0), window=grid_frame, anchor="nw")
        headings = {
            "invoice_date": "Invoice Date",
            "supplier": "Supplier",
            "reference": "Reference",
            "invoice_value": f"Value: {format_usd(data['total_invoices'])}",
            "payment_date": "Payment Date",
            "payment_value": f"Value: {format_usd(data['total_payments'])}",
            "balance": "Balance",
        }
        column_specs = [
            ("invoice_date", 13, "center"),
            ("supplier", 14, "center"),
            ("reference", 18, "center"),
            ("invoice_value", 14, "e"),
            ("payment_date", 13, "center"),
            ("payment_value", 14, "e"),
            ("balance", 14, "e"),
        ]
        border = "#c8d4dc"
        header_bg = "#eef4f7"
        even_bg = "#ffffff"
        odd_bg = "#f8fbfc"
        for column_index, (key, width, anchor) in enumerate(column_specs):
            grid_frame.columnconfigure(column_index, weight=1, minsize=width * 8)
            label = tk.Label(
                grid_frame,
                text=headings[key],
                bg=header_bg,
                fg=TEXT_DARK,
                font=("Segoe UI Semibold", 9),
                anchor=anchor,
                padx=6,
                pady=4,
                relief="solid",
                bd=1,
            )
            label.grid(row=0, column=column_index, sticky="nsew")
        for idx, row in enumerate(data["rows"]):
            row_bg = even_bg if idx % 2 == 0 else odd_bg
            balance_color = NEGATIVE if row["balance"] > 0 else POSITIVE if row["balance"] < 0 else TEXT_DARK
            values = {
                "invoice_date": row["invoice_date"],
                "supplier": row["supplier"],
                "reference": row["reference"],
                "invoice_value": format_usd(row["invoice_value"]) if row["invoice_value"] else "",
                "payment_date": row["payment_date"],
                "payment_value": format_usd(row["payment_value"]) if row["payment_value"] else "",
                "balance": format_usd(row["balance"]),
            }
            for column_index, (key, _width, anchor) in enumerate(column_specs):
                fg = balance_color if key == "balance" else TEXT_DARK
                cell = tk.Label(
                    grid_frame,
                    text=values[key],
                    bg=row_bg,
                    fg=fg,
                    font=("Segoe UI", 9),
                    anchor=anchor,
                    padx=6,
                    pady=3,
                    relief="solid",
                    bd=1,
                )
                cell.grid(row=idx + 1, column=column_index, sticky="nsew")
                cell.configure(highlightbackground=border)

        def _sync_scrollregion(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _sync_canvas_width(event):
            canvas.itemconfigure(canvas_window, width=event.width)

        grid_frame.bind("<Configure>", _sync_scrollregion)
        canvas.bind("<Configure>", _sync_canvas_width)


class RcsDashboardWindow(ttk.Frame):
    def __init__(self, parent, db_path: Path, on_back=None):
        super().__init__(parent, padding=20)
        self.service = RcsDataService(db_path)
        self.on_back = on_back
        self.content_frame = None
        self._build_ui()

    def _build_ui(self):
        self.content_frame = ttk.Frame(self)
        self.content_frame.pack(fill=BOTH, expand=True)
        self.show_home()

    def _set_content(self, widget: ttk.Frame):
        for child in self.content_frame.winfo_children():
            child.destroy()
        widget.master = self.content_frame
        widget.pack(fill=BOTH, expand=True)

    def _go_back_to_main(self):
        if callable(self.on_back):
            self.on_back()

    def show_home(self):
        for child in self.content_frame.winfo_children():
            child.destroy()
        summary = self._get_home_summary()
        wrapper = tk.Frame(self.content_frame, bg=OFF_WHITE)
        wrapper.pack(fill=BOTH, expand=True)

        hero = tk.Frame(wrapper, bg=NAVY, padx=22, pady=18)
        hero.pack(fill="x", pady=(0, 14))
        hero_top = tk.Frame(hero, bg=NAVY)
        hero_top.pack(fill="x")
        ttk.Button(hero_top, text="Back to Main Menu", command=self._go_back_to_main).pack(side="left", padx=(0, 12))
        tk.Label(hero_top, text="RCS Control Center", bg=NAVY, fg="white", font=("Segoe UI Semibold", 19)).pack(side="left")
        tk.Label(
            hero,
            text="Invoices, payments, ledgers, and monthly billing from one compact workspace.",
            bg=NAVY,
            fg="#d8e6f0",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(8, 0))

        stats = tk.Frame(wrapper, bg=OFF_WHITE)
        stats.pack(fill="x", pady=(0, 12))
        stat_items = [
            ("RCSi invoices", f"{summary['invoice_count']} rows", f"{format_usd(summary['invoice_total'])} total"),
            ("Payments", f"{summary['payment_count']} rows", f"{format_usd(summary['payment_total'])} total"),
            ("Monthly invoices", f"{summary['monthly_count']} rows", f"{format_usd(summary['monthly_total'])} billed"),
            ("Marlink -> RCSi", f"{summary['marlink_count']} rows", f"{format_usd(summary['marlink_total'])} cost"),
            ("Iridium SIMs", f"{summary['sim_count']} rows", "Linked number master data"),
            ("Voucher types", f"{summary['voucher_count']} rows", "Reusable cost templates"),
        ]
        for idx, (title, value, detail) in enumerate(stat_items):
            card = tk.Frame(stats, bg="white", highlightbackground="#d7e2ea", highlightthickness=1, padx=14, pady=12)
            card.grid(row=0, column=idx, sticky="nsew", padx=(0, 10) if idx < len(stat_items) - 1 else 0)
            tk.Label(card, text=title, bg="white", fg=MUTED, font=("Segoe UI Semibold", 9)).pack(anchor="w")
            tk.Label(card, text=value, bg="white", fg=TEXT_DARK, font=("Segoe UI Semibold", 16)).pack(anchor="w", pady=(6, 2))
            tk.Label(card, text=detail, bg="white", fg=NAVY, font=("Segoe UI", 10)).pack(anchor="w")
        for idx in range(len(stat_items)):
            stats.columnconfigure(idx, weight=1)

        buttons = [
            ("RCSi -> RCS Kla", "Edit supplier-facing invoice rows and amount paid values.", self.open_rcsi_invoices, "#1f5f99"),
            ("RCS Kla Payments", "Capture incoming payments and track reference flow.", self.open_rcs_payments, "#177e89"),
            ("RCS Kla -> Kla Clients", "Maintain recurring client billing and FX conversion.", self.open_monthly_invoices, "#2c6e49"),
            ("Marlink -> RCSi", "Track Iridium voucher rows with month-grouped totals.", self.open_marlink, "#0f766e"),
            ("RCS Kla Statement", "Review the live Xtralink supplier ledger and balance.", self.open_rcsi_ledger, "#946c00"),
            ("Xtra-Link -> RCS", "Manage Xtralink invoice rows directly.", self.open_xtralink_invoices, "#7a3e65"),
        ]
        grid = tk.Frame(wrapper, bg=OFF_WHITE)
        grid.pack(fill=BOTH, expand=True)
        for idx, (label, description, command, accent) in enumerate(buttons):
            card = tk.Frame(grid, bg="white", highlightbackground="#d7e2ea", highlightthickness=1, padx=16, pady=14)
            card.grid(row=idx // 2, column=idx % 2, sticky="nsew", padx=6, pady=6)
            accent_bar = tk.Frame(card, bg=accent, height=4)
            accent_bar.pack(fill="x", pady=(0, 12))
            tk.Label(card, text=label, bg="white", fg=TEXT_DARK, font=("Segoe UI Semibold", 13)).pack(anchor="w")
            tk.Label(card, text=description, bg="white", fg=MUTED, justify="left", wraplength=340, font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 14))
            ttk.Button(card, text="Open", command=command).pack(anchor="w")
        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)
        for row_idx in range((len(buttons) + 1) // 2):
            grid.rowconfigure(row_idx, weight=1)

        footer = tk.Frame(wrapper, bg=OFF_WHITE)
        footer.pack(fill="x", pady=(8, 0))
        param_shell = tk.Frame(footer, bg="#edf3f8", highlightbackground="#d7e2ea", highlightthickness=1, padx=10, pady=8)
        param_shell.pack(side="right")
        tk.Label(param_shell, text="Parameters", bg="#edf3f8", fg=MUTED, font=("Segoe UI Semibold", 8)).pack(anchor="e")
        ttk.Button(param_shell, text="Open Parameters", command=self.show_parameters).pack(anchor="e", pady=(4, 0))

    def _get_home_summary(self) -> dict:
        with self.service.connect() as conn:
            invoice_row = conn.execute(
                "SELECT COUNT(*) AS qty, COALESCE(SUM(value), 0) AS total FROM rcs_invoices WHERE party = 'RCSi'"
            ).fetchone()
            payment_row = conn.execute(
                "SELECT COUNT(*) AS qty, COALESCE(SUM(value), 0) AS total FROM rcs_payments"
            ).fetchone()
            monthly_row = conn.execute(
                "SELECT COUNT(*) AS qty, COALESCE(SUM(value_usd), 0) AS total FROM rcs_monthly_invoices"
            ).fetchone()
            marlink_row = conn.execute(
                "SELECT COUNT(*) AS qty, COALESCE(SUM(cost), 0) AS total FROM marlink"
            ).fetchone()
            sim_row = conn.execute(
                "SELECT COUNT(*) AS qty FROM iridium_sim"
            ).fetchone()
            voucher_row = conn.execute(
                "SELECT COUNT(*) AS qty FROM iridium_vouchers"
            ).fetchone()
        return {
            "invoice_count": invoice_row["qty"] if invoice_row else 0,
            "invoice_total": invoice_row["total"] if invoice_row else 0.0,
            "payment_count": payment_row["qty"] if payment_row else 0,
            "payment_total": payment_row["total"] if payment_row else 0.0,
            "monthly_count": monthly_row["qty"] if monthly_row else 0,
            "monthly_total": monthly_row["total"] if monthly_row else 0.0,
            "marlink_count": marlink_row["qty"] if marlink_row else 0,
            "marlink_total": marlink_row["total"] if marlink_row else 0.0,
            "sim_count": sim_row["qty"] if sim_row else 0,
            "voucher_count": voucher_row["qty"] if voucher_row else 0,
        }

    def open_rcsi_invoices(self):
        self.show_crud_screen("RCSi -> RCS Kla", {
            "title": "RCSi -> RCS Kla",
            "search_fields": [
                {"key": "date_search", "label": "Date", "type": "date"},
                {"key": "reference_search", "label": "Reference"},
            ],
            "columns": [("invoice_date", "Date"), ("reference", "Reference"), ("value", "Value"), ("party", "Party"), ("comment", "Comment"), ("amount_paid", "Amount Paid")],
            "form_fields": [
                {"key": "invoice_date", "label": "Date", "type": "date", "required": True},
                {"key": "reference", "label": "Reference", "required": True},
                {"key": "value", "label": "Value", "required": True},
                {"key": "party", "label": "Party", "readonly": True, "default": "RCSi"},
                {"key": "comment", "label": "Comment"},
                {"key": "amount_paid", "label": "Amount paid"},
            ],
            "initial_defaults": {"party": "RCSi"},
            "list_method": "list_rcs_invoices",
            "create_method": "create_rcs_invoice",
            "update_method": "update_rcs_invoice",
            "delete_method": "delete_rcs_invoice",
            "fixed_search": {"party": "RCSi"},
            "usd_columns": ["value", "amount_paid"],
            "total_columns": ["value", "amount_paid"],
            "summary_mode": "invoice_amount_paid",
        })

    def open_rcs_payments(self):
        self.show_crud_screen("RCS Kla Payments", {
            "title": "RCS Kla Payments",
            "search_fields": [
                {"key": "date_search", "label": "Date", "type": "date"},
                {"key": "reference_search", "label": "Reference"},
                {"key": "party", "label": "Party", "type": "dropdown", "options": ["All"] + RCS_PAYMENT_PARTIES, "default": "All"},
            ],
            "columns": [("payment_date", "Date"), ("reference", "Reference"), ("value", "Value"), ("party", "Party"), ("comment", "Comment")],
            "form_fields": [
                {"key": "payment_date", "label": "Date", "type": "date", "required": True},
                {"key": "reference", "label": "Reference", "required": True},
                {"key": "value", "label": "Value", "required": True},
                {"key": "party", "label": "Party", "type": "dropdown", "options": RCS_PAYMENT_PARTIES, "required": True},
                {"key": "comment", "label": "Comment"},
            ],
            "list_method": "list_rcs_payments",
            "create_method": "create_rcs_payment",
            "update_method": "update_rcs_payment",
            "delete_method": "delete_rcs_payment",
            "usd_columns": ["value"],
            "total_columns": ["value"],
        })

    def open_monthly_invoices(self):
        self.show_crud_screen("RCS Kla -> Kla Clients", {
            "title": "RCS Kla -> Kla Clients",
            "search_fields": [
                {"key": "date_search", "label": "Date", "type": "date"},
                {"key": "client", "label": "Client", "type": "dropdown", "options": ["All"] + MONTHLY_CLIENTS, "default": "All"},
                {"key": "invoice_number", "label": "Invoice Number"},
            ],
            "columns": [("invoice_date", "Date"), ("client", "Client"), ("value_ugx", "Value (UGX)"), ("exchange_rate", "Exchange Rate"), ("value_usd", "Value ($)"), ("invoice_number", "Invoice Number")],
            "form_fields": [
                {"key": "invoice_date", "label": "Date", "type": "date", "required": True},
                {"key": "client", "label": "Client", "type": "dropdown", "options": MONTHLY_CLIENTS, "required": True},
                {"key": "value_ugx", "label": "Value (UGX)", "required": True},
                {"key": "exchange_rate", "label": "Exchange rate", "required": True, "default": "3900"},
                {"key": "value_usd", "label": "Value ($)", "readonly": True},
                {"key": "invoice_number", "label": "Invoice number", "required": True},
            ],
            "initial_defaults": {"exchange_rate": "3900"},
            "list_method": "list_monthly_invoices",
            "create_method": "create_monthly_invoice",
            "update_method": "update_monthly_invoice",
            "delete_method": "delete_monthly_invoice",
            "ugx_columns": ["value_ugx"],
            "usd_columns": ["value_usd"],
            "total_columns": ["value_ugx", "value_usd"],
            "summary_mode": "monthly_invoice",
            "group_by_month": True,
            "group_date_key": "invoice_date",
            "group_total_columns": ["value_usd", "value_ugx"],
            "group_total_mode": "monthly_invoice",
            "group_column_widths": {
                "invoice_date": 14,
                "client": 22,
                "value_ugx": 16,
                "exchange_rate": 14,
                "value_usd": 14,
                "invoice_number": 18,
            },
        })

    def open_marlink(self):
        self.show_crud_screen("Marlink -> RCSi", {
            "title": "Marlink -> RCSi",
            "search_fields": [
                {"key": "date_search", "label": "Date", "type": "date"},
                {"key": "iridium_number", "label": "Iridium Number"},
                {"key": "voucher_type", "label": "Voucher Type"},
                {"key": "client", "label": "Client"},
                {"key": "marlink_invoice", "label": "Marlink Invoice"},
            ],
            "columns": [
                ("entry_date", "Date"),
                ("iridium_number", "Iridium Number"),
                ("voucher_type", "Voucher Type"),
                ("client", "Client"),
                ("cost", "Cost"),
                ("marlink_invoice", "Marlink Invoice"),
                ("marlink_invoice_value", "Marlink Invoice Value"),
            ],
            "form_fields": [
                {"key": "entry_date", "label": "Date", "type": "date", "required": True},
                {"key": "iridium_number", "label": "Iridium Number", "type": "dropdown", "required": True, "options_method": "sim_msisdns", "combobox_state": "normal"},
                {"key": "voucher_type", "label": "Voucher Type", "type": "dropdown", "required": True, "options_method": "voucher_names"},
                {"key": "client", "label": "Client", "required": True},
                {"key": "cost", "label": "Cost", "required": True},
                {"key": "marlink_invoice", "label": "Marlink Invoice"},
                {"key": "marlink_invoice_value", "label": "Marlink Invoice Value"},
            ],
            "list_method": "list_marlink",
            "create_method": "create_marlink",
            "update_method": "update_marlink",
            "delete_method": "delete_marlink",
            "usd_columns": ["cost", "marlink_invoice_value"],
            "total_columns": ["cost", "marlink_invoice_value"],
            "group_by_month": True,
            "group_date_key": "entry_date",
            "extra_actions": [
                {"label": "Import XLSX", "command": "import_marlink_excel"},
            ],
        })

    def show_parameters(self):
        for child in self.content_frame.winfo_children():
            child.destroy()
        wrapper = tk.Frame(self.content_frame, bg=OFF_WHITE)
        wrapper.pack(fill=BOTH, expand=True)

        header = tk.Frame(wrapper, bg=NAVY, padx=20, pady=16)
        header.pack(fill="x", pady=(0, 12))
        top = tk.Frame(header, bg=NAVY)
        top.pack(fill="x")
        ttk.Button(top, text="Back", command=self.show_home).pack(side="left", padx=(0, 12))
        tk.Label(top, text="Parameters", bg=NAVY, fg="white", font=("Segoe UI Semibold", 18)).pack(side="left")
        tk.Label(
            header,
            text="Rarely used setup tools for Iridium vouchers and SIM master data.",
            bg=NAVY,
            fg="#d8e6f0",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(8, 0))

        body = tk.Frame(wrapper, bg=OFF_WHITE)
        body.pack(fill=BOTH, expand=True)
        buttons = [
            ("Iridium Voucher Types", "Maintain voucher names, units, and default cost values.", self.open_iridium_vouchers, "#4f46e5"),
            ("Iridium SIMs", "Maintain SIM master data and import Excel lists with MSISDN, ICCID, and Client.", self.open_iridium_sims, "#0f4c5c"),
        ]
        for idx, (label, description, command, accent) in enumerate(buttons):
            card = tk.Frame(body, bg="white", highlightbackground="#d7e2ea", highlightthickness=1, padx=16, pady=14)
            card.grid(row=0, column=idx, sticky="nsew", padx=6, pady=6)
            tk.Frame(card, bg=accent, height=4).pack(fill="x", pady=(0, 12))
            tk.Label(card, text=label, bg="white", fg=TEXT_DARK, font=("Segoe UI Semibold", 13)).pack(anchor="w")
            tk.Label(card, text=description, bg="white", fg=MUTED, justify="left", wraplength=340, font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 14))
            ttk.Button(card, text="Open", command=command).pack(anchor="w")
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

    def open_iridium_vouchers(self):
        self.show_crud_screen("Iridium Voucher Types", {
            "title": "Iridium Voucher Types",
            "search_fields": [
                {"key": "name", "label": "Name"},
            ],
            "columns": [
                ("name", "Name"),
                ("units", "Units"),
                ("cost", "Cost"),
            ],
            "form_fields": [
                {"key": "name", "label": "Name", "required": True},
                {"key": "units", "label": "Units", "required": True},
                {"key": "cost", "label": "Cost", "required": True},
            ],
            "list_method": "list_iridium_vouchers",
            "create_method": "create_iridium_voucher",
            "update_method": "update_iridium_voucher",
            "delete_method": "delete_iridium_voucher",
            "usd_columns": ["cost"],
            "total_columns": ["cost"],
        })

    def open_iridium_sims(self):
        self.show_crud_screen("Iridium SIMs", {
            "title": "Iridium SIMs",
            "search_fields": [
                {"key": "msisdn", "label": "MSISDN"},
                {"key": "iccid", "label": "ICCID"},
                {"key": "client", "label": "Client"},
            ],
            "columns": [
                ("msisdn", "MSISDN"),
                ("iccid", "ICCID"),
                ("client", "Client"),
            ],
            "form_fields": [
                {"key": "msisdn", "label": "MSISDN", "required": True},
                {"key": "iccid", "label": "ICCID", "required": True},
                {"key": "client", "label": "Client", "required": True},
            ],
            "list_method": "list_iridium_sims",
            "create_method": "create_iridium_sim",
            "update_method": "update_iridium_sim",
            "delete_method": "delete_iridium_sim",
            "extra_actions": [
                {"label": "Import XLSX", "command": "import_iridium_sims_excel"},
            ],
        })

    def open_rcsi_ledger(self):
        self.show_ledger_screen("RCS Kla Statement", self.service.build_supplier_ledger("Xtra-Link"))

    def open_xtralink_invoices(self):
        self.show_crud_screen("Xtra-Link -> RCS", {
            "title": "Xtra-Link -> RCS",
            "search_fields": [
                {"key": "date_search", "label": "Date", "type": "date"},
                {"key": "reference_search", "label": "Reference"},
            ],
            "columns": [("invoice_date", "Date"), ("reference", "Reference"), ("value", "Value"), ("party", "Party"), ("comment", "Comment"), ("amount_paid", "Amount Paid")],
            "form_fields": [
                {"key": "invoice_date", "label": "Date", "type": "date", "required": True},
                {"key": "reference", "label": "Reference", "required": True},
                {"key": "value", "label": "Value", "required": True},
                {"key": "party", "label": "Party", "readonly": True, "default": "Xtra-Link"},
                {"key": "comment", "label": "Comment"},
                {"key": "amount_paid", "label": "Amount paid"},
            ],
            "initial_defaults": {"party": "Xtra-Link"},
            "list_method": "list_rcs_invoices",
            "create_method": "create_rcs_invoice",
            "update_method": "update_rcs_invoice",
            "delete_method": "delete_rcs_invoice",
            "fixed_search": {"party": "Xtra-Link"},
            "usd_columns": ["value", "amount_paid"],
            "total_columns": ["value", "amount_paid"],
            "summary_mode": "invoice_amount_paid",
        })

    def show_crud_screen(self, title: str, config: dict):
        for child in self.content_frame.winfo_children():
            child.destroy()
        CrudScreen(self.content_frame, title, self.service, config, self.show_home)

    def show_ledger_screen(self, title: str, data: dict):
        for child in self.content_frame.winfo_children():
            child.destroy()
        LedgerScreen(self.content_frame, title, data, self.show_home)
